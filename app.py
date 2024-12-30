from flask import Flask, render_template, request, redirect, url_for, session, jsonify,flash
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import desc
import random
from datetime import datetime
from openpyxl import load_workbook
from werkzeug.utils import secure_filename 
import os
import re
import json

app = Flask(__name__)
UPLOAD_FOLDER = 'static/uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(app.root_path, 'static', 'uploads')
app.config['SECRET_KEY'] = 'your_very_secret_key'
db = SQLAlchemy(app)

class Notification_owner(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject = db.Column(db.String(100), nullable=False)
    message = db.Column(db.String(500), nullable=False)
    viewed = db.Column(db.Boolean, default=False)  
    company_code = db.Column(db.String(20), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.now)


class Notification_organizer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject = db.Column(db.String(100), nullable=False)
    message = db.Column(db.String(500), nullable=False)
    viewed = db.Column(db.Boolean, default=False)  
    company_code = db.Column(db.String(20), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.now)
    
class Notification_admin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_user= db.Column(db.Integer,nullable=False)
    subject = db.Column(db.String(100), nullable=False)
    message = db.Column(db.String(500), nullable=False)
    viewed = db.Column(db.Boolean, default=False)  
    company_code = db.Column(db.String(20), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.now)


class Question_Multiple_Multiple_choice(db.Model):
    question_id = db.Column(db.Integer, primary_key=True, autoincrement=True)  
    exam_id = db.Column(db.Integer, nullable=False)  
    question_text = db.Column(db.Text, nullable=True) 
    audio_path = db.Column(db.String(255), nullable=True) 
    photo_path = db.Column(db.String(255), nullable=True)  
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")  
    bloom_taxonomy = db.Column(db.String(50), default="Analyze") 
    grade = db.Column(db.Integer, default=1)  
    times = db.Column(db.Integer, default=7) 


class Question_sequence(db.Model):
    question_id = db.Column(db.Integer, primary_key=True, autoincrement=True)  
    exam_id = db.Column(db.Integer, nullable=False)  
    question_text = db.Column(db.Text, nullable=True) 
    audio_path = db.Column(db.String(255), nullable=True) 
    photo_path = db.Column(db.String(255), nullable=True)  
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")  
    bloom_taxonomy = db.Column(db.String(50), default="Analyze") 
    grade = db.Column(db.Integer, default=1)  
    times = db.Column(db.Integer, default=7)

class Question_Fill_in_blank(db.Model):
    question_id = db.Column(db.Integer, primary_key=True, autoincrement=True)  
    exam_id = db.Column(db.Integer, nullable=False)  
    question_text = db.Column(db.Text, nullable=True)   
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")  
    bloom_taxonomy = db.Column(db.String(50), default="Analyze") 
    grade = db.Column(db.Integer, default=1)  
    times = db.Column(db.Integer, default=7) 
    answer_text = db.Column(db.Text, nullable=True) 

class Answer_Multiple_Multiple_choice(db.Model):
    answer_id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    question_id = db.Column(db.Integer, nullable=False)  
    answer_text = db.Column(db.Text, nullable=True) 
    audio_path = db.Column(db.String(255), nullable=True)  
    photo_path = db.Column(db.String(255), nullable=True)  
    is_correct = db.Column(db.Boolean, default=False)  

class Answer_sequence(db.Model):
    answer_id = db.Column(db.Integer, primary_key=True, autoincrement=True)  
    question_id = db.Column(db.Integer, nullable=False)  
    answer_text = db.Column(db.Text, nullable=True)  
    index = db.Column(db.Integer, default=False) 

class WordPuzzleGame(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    exam_id = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer, nullable=False)
    company_code = db.Column(db.String(50), nullable=False)
    difficulty = db.Column(db.String(20), nullable=False)
    words = db.Column(db.Text, nullable=False)

class Reverse(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    exam_id = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer, nullable=False)
    company_code = db.Column(db.String(50), nullable=False)
    difficulty = db.Column(db.String(50), nullable=False)
    words = db.Column(db.String(255), nullable=False)

class Quess(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    exam_id = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer, nullable=False)
    company_code = db.Column(db.String(50), nullable=False)
    difficulty = db.Column(db.String(50), nullable=False)
    words = db.Column(db.String(255), nullable=False)


class TraslateAr(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    exam_id = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer, nullable=False)
    company_code = db.Column(db.String(50), nullable=False)
    difficulty = db.Column(db.String(50), nullable=False)
    words = db.Column(db.String(255), nullable=False)
    translate = db.Column(db.String(255), nullable=False)

class Traslateic(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    exam_id = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer, nullable=False)
    company_code = db.Column(db.String(50), nullable=False)
    difficulty = db.Column(db.String(50), nullable=False)
    words = db.Column(db.String(255), nullable=False)
    translate = db.Column(db.String(255), nullable=False)

class Match(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    exam_id = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer, nullable=False)
    company_code = db.Column(db.String(50), nullable=False)
    difficulty = db.Column(db.String(50), nullable=False)
    words = db.Column(db.String(255), nullable=False)

class Chapter(db.Model):
    id = db.Column(db.Integer, primary_key=True, default=lambda: random.randint(100000, 9999999))
    id_chapter=db.Column(db.Integer, nullable=False)
    title = db.Column(db.String(120), nullable=False)  
    image_path = db.Column(db.String(200), nullable=True) 
    created_at = db.Column(db.DateTime, default=datetime.now)  
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    employee_id = db.Column(db.Integer, nullable=False)  

class Exam(db.Model):
    exam_id = db.Column(db.Integer, primary_key=True)
    lesson_id = db.Column(db.Integer, nullable=False)
    exam_name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(255), nullable=False)
    exam_type = db.Column(db.String(50), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow) 

class category(db.Model):
    id = db.Column(db.Integer, primary_key=True, default=lambda: random.randint(100000, 9999999))
    title = db.Column(db.String(120), nullable=False)  
    image_path = db.Column(db.String(200), nullable=True)  
    created_at = db.Column(db.DateTime, default=datetime.now)  
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now) 
    employee_id = db.Column(db.Integer, nullable=False)  

class ManagerEmployee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin=db.Column(db.Integer,nullable=False)
    code = db.Column(db.String(10), nullable=False)
    username = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    birthdate = db.Column(db.Date, nullable=True)
    stage = db.Column(db.String(50), nullable=True)
    university = db.Column(db.String(100), nullable=True)
    major = db.Column(db.String(100), nullable=True)
    phone = db.Column(db.String(15), nullable=True)
    country = db.Column(db.String(100), nullable=True)
    province = db.Column(db.String(100), nullable=True)
    col1 = db.Column(db.String(100), nullable=True)
    col2 = db.Column(db.String(100), nullable=True)
    col3 = db.Column(db.String(100), nullable=True)
    col4 = db.Column(db.String(100), nullable=True)
    col5 = db.Column(db.String(100), nullable=True)
    branch = db.Column(db.String(100), nullable=True) 
    manager_id = db.Column(db.Integer, nullable=True)
    manager_name = db.Column(db.String(50), nullable=True)

class head(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin=db.Column(db.Integer,nullable=False)
    code = db.Column(db.String(10), nullable=False)
    username = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    birthdate = db.Column(db.Date, nullable=True)
    stage = db.Column(db.String(50), nullable=True)
    university = db.Column(db.String(100), nullable=True)
    major = db.Column(db.String(100), nullable=True)
    phone = db.Column(db.String(15), nullable=True)
    country = db.Column(db.String(100), nullable=True)
    province = db.Column(db.String(100), nullable=True)
    col1 = db.Column(db.String(100), nullable=True)
    col2 = db.Column(db.String(100), nullable=True)
    col3 = db.Column(db.String(100), nullable=True)
    col4 = db.Column(db.String(100), nullable=True)
    col5 = db.Column(db.String(100), nullable=True)
    branch = db.Column(db.String(100), nullable=True) 

class Employee6(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin=db.Column(db.Integer,nullable=False)
    code = db.Column(db.String(10), nullable=False)
    username = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    birthdate = db.Column(db.Date, nullable=True)
    stage = db.Column(db.String(50), nullable=True)
    university = db.Column(db.String(100), nullable=True)
    major = db.Column(db.String(100), nullable=True)
    phone = db.Column(db.String(15), nullable=True)
    country = db.Column(db.String(100), nullable=True)
    province = db.Column(db.String(100), nullable=True)
    col1 = db.Column(db.String(100), nullable=True)
    col2 = db.Column(db.String(100), nullable=True)
    col3 = db.Column(db.String(100), nullable=True)
    col4 = db.Column(db.String(100), nullable=True)
    col5 = db.Column(db.String(100), nullable=True)
    branch = db.Column(db.String(100), nullable=True) 
    manager_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=True)
    manager_name = db.Column(db.String(50), nullable=True)
    is_manager = db.Column(db.Boolean, default=False)
    is_head = db.Column(db.String(100), default="fff")

class organizer2(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), nullable=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(50), unique=True, nullable=False)  
    col1 = db.Column(db.String(50), nullable=True)
    col2 = db.Column(db.String(50), nullable=True)
    col3 = db.Column(db.String(50), nullable=True)
    col4 = db.Column(db.String(50), nullable=True)
    col5 = db.Column(db.String(50), nullable=True)
    password = db.Column(db.String(100), nullable=False)

class  administrator2(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(50), unique=True, nullable=False)  
    col1 = db.Column(db.String(50), nullable=True)
    col2 = db.Column(db.String(50), nullable=True)
    col3 = db.Column(db.String(50), nullable=True)
    col4 = db.Column(db.String(50), nullable=True)
    col5 = db.Column(db.String(50), nullable=True)
    code = db.Column(db.String(20), nullable=True)
    password = db.Column(db.String(100), nullable=False)
    
class Employee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    job_number = db.Column(db.String(20), unique=True, nullable=False)
    job_title = db.Column(db.String(50), nullable=False)
    specialization = db.Column(db.String(50), nullable=False)
    company_name = db.Column(db.String(50), nullable=False)
    department = db.Column(db.String(50), nullable=False)
    admin_title = db.Column(db.String(50), nullable=False)
    manager_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=True)

class ColAdministrator(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)
    c1 = db.Column(db.String(50), nullable=False)
    c2 = db.Column(db.String(50), nullable=False)
    c3 = db.Column(db.String(50), nullable=False)
    c4 = db.Column(db.String(50), nullable=False)
    c5 = db.Column(db.String(50), nullable=False)

class ColOrganizer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)
    c1 = db.Column(db.String(50), nullable=False)
    c2 = db.Column(db.String(50), nullable=False)
    c3 = db.Column(db.String(50), nullable=False)
    c4 = db.Column(db.String(50), nullable=False)
    c5 = db.Column(db.String(50), nullable=False)

class ColStudent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), nullable=False)
    id_admin = db.Column(db.Integer, nullable=False)
    c1 = db.Column(db.String(50), nullable=True)
    c2 = db.Column(db.String(50), nullable=True)
    c3 = db.Column(db.String(50), nullable=True)
    c4 = db.Column(db.String(50), nullable=True)
    c5 = db.Column(db.String(50), nullable=True)

class ColEmployee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin = db.Column(db.Integer, nullable=False)
    code = db.Column(db.String(20), nullable=False)
    c1 = db.Column(db.String(50), nullable=True)
    c2 = db.Column(db.String(50), nullable=True)
    c3 = db.Column(db.String(50), nullable=True)
    c4 = db.Column(db.String(50), nullable=True)
    c5 = db.Column(db.String(50), nullable=True)

class Manager(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    job_number = db.Column(db.String(20), unique=True, nullable=False)
    job_title = db.Column(db.String(50), nullable=False)
    specialization = db.Column(db.String(50), nullable=False)
    company_name = db.Column(db.String(50), nullable=False)
    department = db.Column(db.String(50), nullable=False)
    admin_title = db.Column(db.String(50), nullable=False)
    manager_id = db.Column(db.Integer, db.ForeignKey('manager.id'), nullable=True)

class CompanyName(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class specializationEmployee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin=db.Column(db.Integer,nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)    

class Administrative_Title(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin=db.Column(db.Integer,nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False) 

class JobTitle(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin=db.Column(db.Integer,nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class GradeStudent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin = db.Column(db.Integer, nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class DepartmentEmployee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin=db.Column(db.Integer,nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class DepartmentEmployee1(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin=db.Column(db.Integer,nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class DepartmentStudent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin = db.Column(db.Integer, nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class ClassStudent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin = db.Column(db.Integer, nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class yearStudent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin = db.Column(db.Integer, nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class specializationStudent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin = db.Column(db.Integer, nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class ColumnPreference(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), nullable=False)
    column_name = db.Column(db.String(100), nullable=False)
    code=db.Column(db.String(8), nullable=False)
    visible = db.Column(db.Boolean, default=True)

class Personal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(100), nullable=False)
    account_type = db.Column(db.String(50), nullable=False)

class Sector(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    foundation_name = db.Column(db.String(100), nullable=False)
    username = db.Column(db.String(50), nullable=False, unique=False)
    email = db.Column(db.String(100), nullable=False, unique=True)
    code = db.Column(db.String(8), nullable=False, unique=True)
    password = db.Column(db.String(200), nullable=False)
    address = db.Column(db.String(200), nullable=True)
    phone_number = db.Column(db.String(20), nullable=True)
    type = db.Column(db.String(50), nullable=True)
    state = db.Column(db.String(100), nullable=True)
    country = db.Column(db.String(100), nullable=True)
    description = db.Column(db.Text, nullable=True)
    image_path = db.Column(db.String(200), nullable=True)

class Sectorimage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(8), nullable=False, unique=True)
    image_path = db.Column(db.String(200), nullable=True)


class SectorStudent8(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin = db.Column(db.Integer, nullable=False)
    username = db.Column(db.String(150), nullable=False)
    email = db.Column(db.String(150), nullable=False)
    code = db.Column(db.String(8), nullable=False)
    password = db.Column(db.String(200), nullable=False)
    accepted = db.Column(db.Boolean, nullable=False, default=False)
    birthdate = db.Column(db.Date, nullable=True)
    stage = db.Column(db.String(100), nullable=True)
    university = db.Column(db.String(150), nullable=True)
    major = db.Column(db.String(150), nullable=True)
    phone = db.Column(db.String(20), nullable=True)
    country = db.Column(db.String(100), nullable=True)
    province = db.Column(db.String(100), nullable=True)
    col1 = db.Column(db.String(100), nullable=True)
    col2 = db.Column(db.String(100), nullable=True)
    col3 = db.Column(db.String(100), nullable=True)
    col4 = db.Column(db.String(100), nullable=True)
    col5 = db.Column(db.String(100), nullable=True)
    
class mangers(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    foundation_name = db.Column(db.String(100), nullable=False)
    username = db.Column(db.String(50), nullable=False, unique=False)
    email = db.Column(db.String(100), nullable=False, unique=True)
    code = db.Column(db.String(8), nullable=False, unique=True)
    password = db.Column(db.String(200), nullable=False)
    address = db.Column(db.String(200), nullable=True)
    phone_number = db.Column(db.String(20), nullable=True)
    type = db.Column(db.String(50), nullable=True)
    state = db.Column(db.String(100), nullable=True)
    country = db.Column(db.String(100), nullable=True)
    description = db.Column(db.Text, nullable=True)

class Education(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    foundation_name = db.Column(db.String(100), nullable=False)
    username = db.Column(db.String(50), nullable=False, unique=False)
    email = db.Column(db.String(100), nullable=False, unique=True)
    password = db.Column(db.String(200), nullable=False)

class Questionfor_t_or_f(db.Model):
    question_id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    exam_id = db.Column(db.Integer,  nullable=False)
    audio_path = db.Column(db.String(255), default="")
    photo_path = db.Column(db.String(255), default="")
    question_name = db.Column(db.String(255), default="السؤال الأول")
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")
    bloom_taxonomy = db.Column(db.String(50), default="Analyze")
    grade = db.Column(db.Integer, default=1)
    times = db.Column(db.Integer, default=7)
    correct_answer = db.Column(db.Boolean, default=True)

with app.app_context():
    #db.drop_all()
    db.create_all()
@app.route('/update_preferences', methods=['GET', 'POST'])
def update_preferences():
    if request.method == 'POST':
        selected_columns = request.form.getlist('columns')
        email = "user@example.com"

        ColumnPreference.query.filter_by(email=email).delete()

        for column in ['id', 'username', 'email', 'password', 'account_type']:
            preference = ColumnPreference(
                email=email,
                column_name=column,
                visible=(column in selected_columns)
            )
            db.session.add(preference)
        db.session.commit()

        return redirect(url_for('success', username="user"))

    return render_template('update_preferences.html')

@app.route('/signup_education', methods=['GET', 'POST'])
def signup_education():
    if request.method == 'POST' and request.form.get('foundation_name')!=None:
        foundation_name = request.form.get('foundation_name')
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        existing_Education = Education.query.filter_by(email=email).first()
        if existing_Education:
            return render_template('signup_education.html', error="Email already exists. Please choose a different one.")

        if password != confirm_password:
            return render_template('signup_education.html', error="Passwords do not match.")

        new_Education = Education(
            foundation_name=foundation_name,
            username=username,
            email=email,
            password=password 
        )
        db.session.add(new_Education)
        db.session.commit()

        return redirect(url_for('success', username=username))

    return render_template('signup_education.html')   

@app.route('/access', methods=['GET', 'POST'])
def access():
    return render_template('access.html')  

@app.route('/accept', methods=['GET', 'POST'])
def accept():
    return render_template('accept.html')  

@app.route('/edit_question_true_or_false', methods=['GET', 'POST'])
def edit_question_true_or_false():
    exam_id = request.args.get('exam_id')
    questions = Questionfor_t_or_f.query.filter_by(exam_id=exam_id).all()
    return render_template('edit_question_true_or_false.html',exam_id=exam_id,questions=questions)  

@app.route('/add_question_T_or_F/<int:exam_id>', methods=['POST'])
def add_question_T_or_F(exam_id):
    question = Questionfor_t_or_f(
        exam_id=exam_id,
        question_name="السؤال الأول",
        difficulty="Easy",
        importance="Necessary",
        bloom_taxonomy="Analyze",
        grade=1,
        times=7,
        correct_answer=True
    )
    db.session.add(question)
    db.session.commit()

    return redirect(url_for('edit_question_true_or_false', exam_id=exam_id))

@app.route('/edit_question_Multiple_Multiple', methods=['GET', 'POST'])
def edit_question_Multiple_Multiple():
    exam_id = request.args.get('exam_id')
    questions = Question_Multiple_Multiple_choice.query.filter_by(exam_id=exam_id).all()
    return render_template('edit_question_Multiple_Multiple.html',exam_id=exam_id,questions=questions) 

@app.route('/edit_fill_in_blank', methods=['GET', 'POST'])
def edit_fill_in_blank():
    exam_id = request.args.get('exam_id')
    questions = Question_Fill_in_blank.query.filter_by(exam_id=exam_id).all()
    return render_template('edit_fill_in_blank.html',exam_id=exam_id,questions=questions) 


@app.route('/add_fill_in_blank/<int:exam_id>', methods=['POST'])
def add_fill_in_blank(exam_id):
    question = Question_Fill_in_blank(
        exam_id=exam_id,
        question_text="السؤال الافتراضي",
        difficulty="Easy",
        importance="Necessary",
        bloom_taxonomy="Analyze",
        grade=1,
        times=7,
        answer_text="الاجابه"
    )
    db.session.add(question)
    db.session.commit()
    
    return redirect(url_for('edit_fill_in_blank', exam_id=exam_id))


@app.route('/fill_in_blank', methods=['GET', 'POST'])
def fill_in_blank():
    question_id = request.args.get('question_id', type=int)
    question = Question_Fill_in_blank.query.get_or_404(question_id)

    if request.method == 'POST':
        question.question_text = request.form['question_text']
        question.difficulty = request.form['difficulty']
        question.importance = request.form['importance']
        question.bloom_taxonomy = request.form['bloom_taxonomy']
        question.grade = request.form['grade']
        question.times = request.form['times']
        question.answer_text = request.form['answer']
      
        try:
            db.session.commit()
        except Exception as e:
            print(f"Error saving data: {e}")

    return render_template('fill_in_blank.html', question=question)

@app.route('/edit_question_Multiple_single', methods=['GET', 'POST'])
def edit_question_Multiple_single():
    exam_id = request.args.get('exam_id')
    questions = Question_Multiple_Multiple_choice.query.filter_by(exam_id=exam_id).all()
    return render_template('edit_question_Multiple_single.html',exam_id=exam_id,questions=questions)   

@app.route('/edit_question_sequence', methods=['GET', 'POST'])
def edit_question_sequence():
    exam_id = request.args.get('exam_id')
    questions = Question_sequence.query.filter_by(exam_id=exam_id).all()
    return render_template('edit_question_sequence.html',exam_id=exam_id,questions=questions)  

@app.route('/add_question_sequence/<int:exam_id>', methods=['POST'])
def add_question_sequence(exam_id):
    question = Question_sequence(
        exam_id=exam_id,
        question_text="السؤال الافتراضي",
        difficulty="Easy",
        importance="Necessary",
        bloom_taxonomy="Analyze",
        grade=1,
        times=7
    )
    db.session.add(question)
    db.session.commit()
    answers = [
        Answer_sequence(
            question_id=question.question_id,
            answer_text=f"الإجابة الافتراضية ",
            index=i 
        )
        for i in range(4)  
    ]
    db.session.add_all(answers)
    db.session.commit()

    return redirect(url_for('edit_question_sequence', exam_id=exam_id))

@app.route('/sequence/<int:question_id>', methods=['GET', 'POST'])
def sequence(question_id):
    # جلب السؤال من قاعدة البيانات
    question = Question_sequence.query.get(question_id)
    
    if request.method == 'GET':
        # جلب الإجابات المرتبطة بالسؤال
        answers = Answer_sequence.query.filter_by(question_id=question_id).all()
        return render_template('sequence.html', question=question, answers=answers)

    if request.method == 'POST':
        # تحديث بيانات السؤال من الـ POST
        question.question_text = request.form.get('question_text')
        question.difficulty = request.form.get('difficulty')
        question.importance = request.form.get('importance')
        question.bloom_taxonomy = request.form.get('bloom_taxonomy')
        question.grade = request.form.get('grade')
        question.times = request.form.get('times')

        # معالجة الصورة
        if 'photo' in request.files:
            photo = request.files['photo']
            if photo and allowed_file1(photo.filename):
                if question.photo_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path))
                    except FileNotFoundError:
                        pass
                filename = f"s_{question_id}_{secure_filename(photo.filename).rsplit('.', 1)[0]}.png"
                try:
                    photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.photo_path = filename
                except Exception as e:
                    print(f"Error saving photo: {e}")

        # معالجة الصوت
        if 'audio' in request.files:
            audio = request.files['audio']
            if audio and allowed_file1(audio.filename):
                if question.audio_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path))
                    except FileNotFoundError:
                        pass
                filename = f"s_{question_id}_{secure_filename(audio.filename)}"
                try:
                    audio.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.audio_path = filename
                except Exception as e:
                    print(f"Error saving audio: {e}")

        answers = Answer_sequence.query.filter_by(question_id=question_id).all()
        for i, answer in enumerate(answers):
            # تحديث النص
            answer.answer_text = request.form.get(f'answer_text_{i}')
            
            # تحديث الترتيب بناءً على القيمة المدخلة
            order_value = request.form.get(f'order_{i}')
            if order_value is not None and order_value.isdigit():
                answer.index = int(order_value)
            else:
                answer.index = 0  # تعيين القيمة الافتراضية إذا لم يتم إدخال ترتيب صحيح

        # حفظ التعديلات
        db.session.commit()
        return render_template('sequence.html', question=question, answers=answers)


@app.route('/add_question_Multiple_single_choice/<int:exam_id>', methods=['POST'])
def add_question_Multiple_single_choice(exam_id):
    question = Question_Multiple_Multiple_choice(
        exam_id=exam_id,
        question_text="السؤال الافتراضي",
        difficulty="Easy",
        importance="Necessary",
        bloom_taxonomy="Analyze",
        grade=1,
        times=7
    )
    db.session.add(question)
    db.session.commit()
    answers = [
        Answer_Multiple_Multiple_choice(
            question_id=question.question_id,
            answer_text=f"الإجابة الافتراضية {i+1}",
            is_correct=(i == 0)  
        )
        for i in range(4)  
    ]
    db.session.add_all(answers)
    db.session.commit()

    return redirect(url_for('edit_question_Multiple_single', exam_id=exam_id))




@app.route('/mcq_multiple_single/<int:question_id>', methods=['GET', 'POST'])
def mcq_multiple_single(question_id):
  
    question = Question_Multiple_Multiple_choice.query.get(question_id)
    
    if request.method == 'GET':
        answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=question_id).all()
        return render_template('mcq_multiple_single.html', question=question, answers=answers)

    if request.method == 'POST':
        question.question_text = request.form.get('question_text')
        question.difficulty = request.form.get('difficulty')
        question.importance = request.form.get('importance')
        question.bloom_taxonomy = request.form.get('bloom_taxonomy')
        question.grade = request.form.get('grade')
        question.times = request.form.get('times')

        answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=question_id).all()
        correct_index = int(request.form.get('correct'))
        for i, answer in enumerate(answers):
            answer.answer_text = request.form.get(f'answer_text_{i}')
            answer.is_correct = (i == correct_index)  
            if 'photo_' + str(i) in request.files:
                photo = request.files['photo_' + str(i)]
                if photo and allowed_file1(photo.filename):
                   
                    if answer.photo_path:
                        try:
                            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], answer.photo_path))
                        except FileNotFoundError:
                            pass

                    filename = f"M_S_{question_id}_answer_{i}_{secure_filename(photo.filename).rsplit('.', 1)[0]}.png"
                    try:
                        photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                        answer.photo_path = filename
                    except Exception as e:
                        print(f"Error saving photo: {e}")

         
            if 'audio_' + str(i) in request.files:
                audio = request.files['audio_' + str(i)]
                if audio and allowed_file1(audio.filename):
                  
                    if answer.audio_path:
                        try:
                            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], answer.audio_path))
                        except FileNotFoundError:
                            pass

                   
                    filename = f"M_S_{question_id}_answer_{i}_{secure_filename(audio.filename)}"
                    try:
                        audio.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                        answer.audio_path = filename
                    except Exception as e:
                        print(f"Error saving audio: {e}")


        if 'photo' in request.files:
            photo = request.files['photo']
            if photo and allowed_file1(photo.filename):
      
                if question.photo_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path))
                    except FileNotFoundError:
                        pass

                filename = f"M_S_{question_id}_{secure_filename(photo.filename).rsplit('.', 1)[0]}.png"
                try:
                    photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.photo_path = filename
                except Exception as e:
                    print(f"Error saving photo: {e}")

        # معالجة رفع الصوت
        if 'audio' in request.files:
            audio = request.files['audio']
            if audio and allowed_file1(audio.filename):
          
                if question.audio_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path))
                    except FileNotFoundError:
                        pass

                
                filename = f"M_S_{question_id}_{secure_filename(audio.filename)}"
                try:
                    audio.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.audio_path = filename
                except Exception as e:
                    print(f"Error saving audio: {e}")


        db.session.commit()
        return render_template('mcq_multiple_single.html', question=question, answers=answers)


@app.route('/mcq_multiple_ans/<int:question_id>', methods=['GET', 'POST'])
def mcq_multiple_ans(question_id):
  
    question = Question_Multiple_Multiple_choice.query.get(question_id)
    
    if request.method == 'GET':
        answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=question_id).all()
        return render_template('mcq_multiple_ans.html', question=question, answers=answers)

    if request.method == 'POST':
        question.question_text = request.form.get('question_text')
        question.difficulty = request.form.get('difficulty')
        question.importance = request.form.get('importance')
        question.bloom_taxonomy = request.form.get('bloom_taxonomy')
        question.grade = request.form.get('grade')
        question.times = request.form.get('times')

        answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=question_id).all()
        for i, answer in enumerate(answers):
            answer.answer_text = request.form.get(f'answer_text_{i}')
            answer.is_correct = f'correct_{i}' in request.form
            if 'photo_' + str(i) in request.files:
                photo = request.files['photo_' + str(i)]
                if photo and allowed_file1(photo.filename):
                   
                    if answer.photo_path:
                        try:
                            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], answer.photo_path))
                        except FileNotFoundError:
                            pass

                    filename = f"M_M_{question_id}_answer_{i}_{secure_filename(photo.filename).rsplit('.', 1)[0]}.png"
                    try:
                        photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                        answer.photo_path = filename
                    except Exception as e:
                        print(f"Error saving photo: {e}")

         
            if 'audio_' + str(i) in request.files:
                audio = request.files['audio_' + str(i)]
                if audio and allowed_file1(audio.filename):
                  
                    if answer.audio_path:
                        try:
                            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], answer.audio_path))
                        except FileNotFoundError:
                            pass

                    # تحديد الاسم الجديد للصوت مع الحفاظ على الامتداد الأصلي
                    filename = f"M_M_{question_id}_answer_{i}_{secure_filename(audio.filename)}"
                    try:
                        audio.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                        answer.audio_path = filename
                    except Exception as e:
                        print(f"Error saving audio: {e}")


        if 'photo' in request.files:
            photo = request.files['photo']
            if photo and allowed_file1(photo.filename):
                # حذف الصورة القديمة إذا كانت موجودة
                if question.photo_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path))
                    except FileNotFoundError:
                        pass

                # تحديد الاسم الجديد للصورة مع تغيير الامتداد إلى .png
                filename = f"M_M_{question_id}_{secure_filename(photo.filename).rsplit('.', 1)[0]}.png"
                try:
                    photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.photo_path = filename
                except Exception as e:
                    print(f"Error saving photo: {e}")

        # معالجة رفع الصوت
        if 'audio' in request.files:
            audio = request.files['audio']
            if audio and allowed_file1(audio.filename):
                # حذف الصوت القديم إذا كان موجودًا
                if question.audio_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path))
                    except FileNotFoundError:
                        pass

                # تحديد الاسم الجديد للصوت مع الحفاظ على الامتداد الأصلي
                filename = f"M_M_{question_id}_{secure_filename(audio.filename)}"
                try:
                    audio.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.audio_path = filename
                except Exception as e:
                    print(f"Error saving audio: {e}")


        db.session.commit()
       

        return render_template('mcq_multiple_ans.html', question=question, answers=answers)


@app.route('/add_question_Multiple_Multiple_choice/<int:exam_id>', methods=['POST'])
def add_question_Multiple_Multiple_choice(exam_id):
    question = Question_Multiple_Multiple_choice(
        exam_id=exam_id,
        question_text="السؤال الافتراضي",
        difficulty="Easy",
        importance="Necessary",
        bloom_taxonomy="Analyze",
        grade=1,
        times=7
    )
    db.session.add(question)
    db.session.commit()
    answers = [
        Answer_Multiple_Multiple_choice(
            question_id=question.question_id,
            answer_text=f"الإجابة الافتراضية {i+1}",
            is_correct=(i == 0)  
        )
        for i in range(4)  
    ]
    db.session.add_all(answers)
    db.session.commit()

    return redirect(url_for('edit_question_Multiple_Multiple', exam_id=exam_id))


@app.route('/edit_exams', methods=['GET', 'POST'])
def edit_exams():
    exams = Exam.query.all()
    return render_template('edit_exams.html',exams=exams) 

@app.route('/Word_Puzzle', methods=['GET', 'POST'])
def Word_Puzzle():
    exam_id = request.args.get('exam_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')

    if not user_id or not company_code:
        return redirect(url_for('login'))  

    if request.method == 'POST':
        difficulty = request.form.get('col1')
        Word_Puzzle = request.form.get('Word_Puzzle')
        uploaded_file = request.files.get('Word_Puzzle_file')

        # إدخال كلمة من حقل النص
        if Word_Puzzle:
            try:
                # التحقق من عدم وجود الكلمة مسبقًا
                existing_word = WordPuzzleGame.query.filter_by(
                    exam_id=exam_id, words=Word_Puzzle
                ).first()
                if not existing_word:
                    new_word = WordPuzzleGame(
                        exam_id=exam_id,
                        user_id=user_id,
                        company_code=company_code,
                        difficulty=difficulty,
                        words=Word_Puzzle
                    )
                    db.session.add(new_word)
                    db.session.commit()
                else:
                    print("Word already exists for this exam.")
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        # رفع ملف Excel
        if uploaded_file and uploaded_file.filename != '':
            filename = secure_filename(uploaded_file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            uploaded_file.save(filepath)

            try:
                workbook = load_workbook(filepath)
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2):
                    word = row[0].value
                    if word: 
                        existing_word = WordPuzzleGame.query.filter_by(
                            exam_id=exam_id, words=word
                        ).first()
                        if not existing_word:
                            new_word = WordPuzzleGame(
                                exam_id=exam_id,
                                user_id=user_id,
                                company_code=company_code,
                                difficulty=difficulty,
                                words=word
                            )
                            db.session.add(new_word)
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")
            finally:
                os.remove(filepath)

    games = WordPuzzleGame.query.filter_by(exam_id=exam_id).all()
    last_difficulty = WordPuzzleGame.query.filter_by(exam_id=exam_id).order_by(WordPuzzleGame.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy' 

    return render_template('games/Word_Puzzle.html', games=games, exam_id=exam_id, current_difficulty=current_difficulty)


@app.route('/match', methods=['GET', 'POST'])
def match():
    exam_id = request.args.get('exam_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')

    if not user_id or not company_code:
        return redirect(url_for('login'))  

    if request.method == 'POST':
        difficulty = request.form.get('col1')
        Word_Puzzle = request.form.get('Word_Puzzle')
        uploaded_file = request.files.get('Word_Puzzle_file')

        if Word_Puzzle:
            try:
                existing_word = Match.query.filter_by(
                    exam_id=exam_id, words=Word_Puzzle
                ).first()
                if not existing_word:
                    new_word = Match(
                        exam_id=exam_id,
                        user_id=user_id,
                        company_code=company_code,
                        difficulty=difficulty,
                        words=Word_Puzzle
                    )
                    db.session.add(new_word)
                    db.session.commit()
                else:
                    print("Word already exists for this exam.")
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        if uploaded_file and uploaded_file.filename != '':
            filename = secure_filename(uploaded_file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            uploaded_file.save(filepath)

            try:
                workbook = load_workbook(filepath)
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2):
                    word = row[0].value
                    if word: 
                        existing_word = Match.query.filter_by(
                            exam_id=exam_id, words=word
                        ).first()
                        if not existing_word:
                            new_word = Match(
                                exam_id=exam_id,
                                user_id=user_id,
                                company_code=company_code,
                                difficulty=difficulty,
                                words=word
                            )
                            db.session.add(new_word)
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")
            finally:
                os.remove(filepath)

    games = Match.query.filter_by(exam_id=exam_id).all()
    last_difficulty = Match.query.filter_by(exam_id=exam_id).order_by(Match.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy' 

    return render_template('games/match.html', games=games, exam_id=exam_id, current_difficulty=current_difficulty)

@app.route('/reverse', methods=['GET', 'POST'])
def reverse():
    exam_id = request.args.get('exam_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')

    if not user_id or not company_code:
        return redirect(url_for('login'))  

    if request.method == 'POST':
        difficulty = request.form.get('col1')
        Word_Puzzle = request.form.get('Word_Puzzle')
        uploaded_file = request.files.get('Word_Puzzle_file')

        if Word_Puzzle:
            try:
                existing_word = Reverse.query.filter_by(
                    exam_id=exam_id, words=Word_Puzzle
                ).first()
                if not existing_word:
                    new_word = Reverse(
                        exam_id=exam_id,
                        user_id=user_id,
                        company_code=company_code,
                        difficulty=difficulty,
                        words=Word_Puzzle
                    )
                    db.session.add(new_word)
                    db.session.commit()
                else:
                    print("Word already exists for this exam.")
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        if uploaded_file and uploaded_file.filename != '':
            filename = secure_filename(uploaded_file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            uploaded_file.save(filepath)

            try:
                workbook = load_workbook(filepath)
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2):
                    word = row[0].value
                    if word: 
                        existing_word = Reverse.query.filter_by(
                            exam_id=exam_id, words=word
                        ).first()
                        if not existing_word:
                            new_word = Reverse(
                                exam_id=exam_id,
                                user_id=user_id,
                                company_code=company_code,
                                difficulty=difficulty,
                                words=word
                            )
                            db.session.add(new_word)
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")
            finally:
                os.remove(filepath)

    games = Reverse.query.filter_by(exam_id=exam_id).all()
    last_difficulty = Reverse.query.filter_by(exam_id=exam_id).order_by(Reverse.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy' 

    return render_template('games/reverse.html', games=games, exam_id=exam_id, current_difficulty=current_difficulty)


@app.route('/quess', methods=['GET', 'POST'])
def quess():
    exam_id = request.args.get('exam_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')

    if not user_id or not company_code:
        return redirect(url_for('login'))  

    if request.method == 'POST':
        difficulty = request.form.get('col1')
        Word_Puzzle = request.form.get('Word_Puzzle')
        uploaded_file = request.files.get('Word_Puzzle_file')

        if Word_Puzzle:
            try:
                existing_word = Quess.query.filter_by(
                    exam_id=exam_id, words=Word_Puzzle
                ).first()
                if not existing_word:
                    new_word = Quess(
                        exam_id=exam_id,
                        user_id=user_id,
                        company_code=company_code,
                        difficulty=difficulty,
                        words=Word_Puzzle
                    )
                    db.session.add(new_word)
                    db.session.commit()
                else:
                    print("Word already exists for this exam.")
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        if uploaded_file and uploaded_file.filename != '':
            filename = secure_filename(uploaded_file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            uploaded_file.save(filepath)

            try:
                workbook = load_workbook(filepath)
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2):
                    word = row[0].value
                    if word: 
                        existing_word = Quess.query.filter_by(
                            exam_id=exam_id, words=word
                        ).first()
                        if not existing_word:
                            new_word = Quess(
                                exam_id=exam_id,
                                user_id=user_id,
                                company_code=company_code,
                                difficulty=difficulty,
                                words=word
                            )
                            db.session.add(new_word)
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")
            finally:
                os.remove(filepath)

    games = Quess.query.filter_by(exam_id=exam_id).all()
    last_difficulty = Quess.query.filter_by(exam_id=exam_id).order_by(Quess.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy' 

    return render_template('games/quess.html', games=games, exam_id=exam_id, current_difficulty=current_difficulty)

@app.route('/traslate', methods=['GET', 'POST'])
def traslate():
    exam_id = request.args.get('exam_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')

    if not user_id or not company_code:
        return redirect(url_for('login'))  

    if request.method == 'POST':
        difficulty = request.form.get('col1')
        Word_Puzzle = request.form.get('Word_Puzzle')
        translate = request.form.get('translate')

        if Word_Puzzle:
            try:
                existing_word = TraslateAr.query.filter_by(
                    exam_id=exam_id, words=Word_Puzzle
                ).first()
                if not existing_word:
                    new_word = TraslateAr(
                        exam_id=exam_id,
                        user_id=user_id,
                        company_code=company_code,
                        difficulty=difficulty,
                        translate=translate,
        
                        words=Word_Puzzle
                    )
                    db.session.add(new_word)
                    db.session.commit()
                else:
                    print("Word already exists for this exam.")
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

    
    games = TraslateAr.query.filter_by(exam_id=exam_id).all()
    last_difficulty = TraslateAr.query.filter_by(exam_id=exam_id).order_by(TraslateAr.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy' 

    return render_template('games/traslate.html', games=games, exam_id=exam_id, current_difficulty=current_difficulty)


@app.route('/traslate_icon', methods=['GET', 'POST'])
def traslate_icon():
    exam_id = request.args.get('exam_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')

    if not user_id or not company_code:
        return redirect(url_for('login'))  

    if request.method == 'POST':
        difficulty = request.form.get('col1')
        Word_Puzzle = request.form.get('Word_Puzzle')
        translate = request.form.get('translate')

        if Word_Puzzle:
            try:
                existing_word = Traslateic.query.filter_by(
                    exam_id=exam_id, words=Word_Puzzle
                ).first()
                if not existing_word:
                    new_word = Traslateic(
                        exam_id=exam_id,
                        user_id=user_id,
                        company_code=company_code,
                        difficulty=difficulty,
                        translate=translate,
        
                        words=Word_Puzzle
                    )
                    db.session.add(new_word)
                    db.session.commit()
                else:
                    print("Word already exists for this exam.")
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

    
    games = Traslateic.query.filter_by(exam_id=exam_id).all()
    last_difficulty = Traslateic.query.filter_by(exam_id=exam_id).order_by(Traslateic.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy' 

    return render_template('games/traslate_icon.html', games=games, exam_id=exam_id, current_difficulty=current_difficulty)

@app.route('/word_game', methods=['GET', 'POST'])
def word_game():
    exam_id = request.args.get('exam_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')
    games = WordPuzzleGame.query.filter_by(exam_id=exam_id).all()
    words = WordPuzzleGame.query.filter_by(exam_id=exam_id).all()
    word_list = [word.words for word in words]

    last_difficulty = WordPuzzleGame.query.filter_by(exam_id=exam_id).order_by(WordPuzzleGame.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy'

    return render_template('games/word_game.html', 
                           games=games, 
                           exam_id=exam_id, 
                           current_difficulty=current_difficulty, 
                           word_list=word_list)

@app.route('/reverse_game', methods=['GET', 'POST'])
def reverse_game():
    exam_id = request.args.get('exam_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')
    games = Reverse.query.filter_by(exam_id=exam_id).all()
    words = Reverse.query.filter_by(exam_id=exam_id).all()
    word_list = [word.words for word in words]
    last_difficulty = Reverse.query.filter_by(exam_id=exam_id).order_by(Reverse.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy'

    return render_template('games/reverse_game.html', 
                           games=games, 
                           exam_id=exam_id, 
                           current_difficulty=current_difficulty, 
                           word_list=word_list)

@app.route('/quess_game', methods=['GET', 'POST'])
def quess_game():
    exam_id = request.args.get('exam_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')
    games = Quess.query.filter_by(exam_id=exam_id).all()
    words = Quess.query.filter_by(exam_id=exam_id).all()
    word_list = [word.words for word in words]
    last_difficulty = Quess.query.filter_by(exam_id=exam_id).order_by(Quess.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy'

    return render_template('games/quess_game.html', 
                           games=games, 
                           exam_id=exam_id, 
                           current_difficulty=current_difficulty, 
                           word_list=word_list)

@app.route('/match_game', methods=['GET', 'POST'])
def match_game():
    exam_id = request.args.get('exam_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')
    games = Match.query.filter_by(exam_id=exam_id).all()
    words = Match.query.filter_by(exam_id=exam_id).all()
    word_list = [word.words for word in words]
    print(word_list)
    print(word_list)
    print(word_list)
    last_difficulty = Match.query.filter_by(exam_id=exam_id).order_by(Match.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy'

    return render_template('games/match_game.html', 
                           games=games, 
                           exam_id=exam_id, 
                           current_difficulty=current_difficulty, 
                           word_list=word_list)

@app.route('/traslate_game', methods=['GET', 'POST'])
def traslate_game():
    exam_id = request.args.get('exam_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')
    games = TraslateAr.query.filter_by(exam_id=exam_id).all()
    words = TraslateAr.query.filter_by(exam_id=exam_id).all()
    word_list = [word.words for word in words]
    translate_list = [word.translate for word in words]
    last_difficulty = TraslateAr.query.filter_by(exam_id=exam_id).order_by(TraslateAr.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy'

    return render_template('games/traslate_game.html', 
                           games=games, 
                           exam_id=exam_id, 
                           current_difficulty=current_difficulty, 
                           word_list=word_list,
                           translate_list=translate_list)

@app.route('/traslate_icon_game', methods=['GET', 'POST'])
def traslate_icon_game():
    exam_id = request.args.get('exam_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')
    games = Traslateic.query.filter_by(exam_id=exam_id).all()
    words = Traslateic.query.filter_by(exam_id=exam_id).all()
    word_list = [word.words for word in words]
    translate_list = [word.translate for word in words]
    last_difficulty = Traslateic.query.filter_by(exam_id=exam_id).order_by(Traslateic.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy'

    return render_template('games/traslate_icon_game.html', 
                           games=games, 
                           exam_id=exam_id, 
                           current_difficulty=current_difficulty, 
                           word_list=word_list,
                           translate_list=translate_list)


@app.route('/delete_word/<int:word_id>', methods=['GET'])
def delete_word(word_id):
    word = WordPuzzleGame.query.get_or_404(word_id)
    try:
        db.session.delete(word)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
    return redirect(url_for('Word_Puzzle', exam_id=word.exam_id))

@app.route('/delete_match/<int:word_id>', methods=['GET'])
def delete_match(word_id):
    word = Match.query.get_or_404(word_id)
    try:
        db.session.delete(word)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
    return redirect(url_for('match', exam_id=word.exam_id))


@app.route('/delete_reverse/<int:word_id>', methods=['GET'])
def delete_reverse(word_id):
    word = Reverse.query.get_or_404(word_id)
    try:
        db.session.delete(word)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
    return redirect(url_for('reverse', exam_id=word.exam_id))

@app.route('/delete_quess/<int:word_id>', methods=['GET'])
def delete_quess(word_id):
    word = Quess.query.get_or_404(word_id)
    try:
        db.session.delete(word)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
    return redirect(url_for('quess', exam_id=word.exam_id))

@app.route('/delete_traslate/<int:word_id>', methods=['GET'])
def delete_traslate(word_id):
    word = TraslateAr.query.get_or_404(word_id)
    try:
        db.session.delete(word)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
    return redirect(url_for('traslate', exam_id=word.exam_id))

@app.route('/delete_traslate_icon/<int:word_id>', methods=['GET'])
def delete_traslate_icon(word_id):
    word = Traslateic.query.get_or_404(word_id)
    try:
        db.session.delete(word)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
    return redirect(url_for('traslate_icon', exam_id=word.exam_id))


@app.route('/redirect_exam/<int:exam_type>/<int:exam_id>', methods=['GET', 'POST'])
def redirect_exam(exam_type, exam_id):
    
    exam = Exam.query.get_or_404(exam_id)

    if exam_type == 1:  
        return redirect(url_for('edit_question_Multiple_Multiple', exam_id=exam_id))
    elif exam_type == 2:  
        return redirect(url_for('edit_question_true_or_false', exam_id=exam_id))
    elif exam_type == 3:  
        return redirect(url_for('edit_fill_in_blank', exam_id=exam_id))
    elif exam_type == 4:  
        return redirect(url_for('edit_question_sequence', exam_id=exam_id))
    elif exam_type == 6:  
        return redirect(url_for('edit_question_Multiple_single', exam_id=exam_id))
    elif exam_type == 12:  
        return redirect(url_for('Word_Puzzle', exam_id=exam_id))
    elif exam_type == 13:  
        return redirect(url_for('match', exam_id=exam_id))
    elif exam_type == 14:  
        return redirect(url_for('reverse', exam_id=exam_id))
    elif exam_type == 15:  
        return redirect(url_for('traslate', exam_id=exam_id))
    elif exam_type == 16:  
        return redirect(url_for('traslate_icon', exam_id=exam_id))
    elif exam_type == 17:
        return redirect(url_for('quess', exam_id=exam_id))
    else:
        return redirect(url_for('general_exam_page', exam_id=exam_id))
    
@app.route('/delete_exam/<int:exam_id>', methods=['POST'])
def delete_exam(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    
    db.session.delete(exam)
    db.session.commit()
    
    return redirect(url_for('edit_exams'))

@app.route('/questions', methods=['GET', 'POST'])
def questions():
    lesson_id = request.args.get('id')
    return render_template('questions.html',lesson_id=lesson_id)  

@app.route('/submit_exam/<int:lesson_id>', methods=['POST'])
def submit_exam(lesson_id):
    
    exam_name = request.form['exam_name']
    description = request.form['description']
    exam_type = request.form['exam_type']

    new_exam = Exam(
        lesson_id=lesson_id,
        exam_name=exam_name,
        description=description,
        exam_type=exam_type
    )
    db.session.add(new_exam)
    db.session.commit()

    return redirect(url_for('edit_exams')) 

@app.route('/exam_cards', methods=['GET', 'POST'])
def exam_cards():
    employee_id = session.get('employee_id')
    sector_code = session.get('sector_code')

    if not sector_code or not employee_id:
        return redirect(url_for('access'))

    lessons = category.query.filter_by(employee_id=employee_id).all()
    lessons_with_counts = []
    for lesson in lessons:
        lesson_count = Chapter.query.filter_by(id_chapter=lesson.id).count()
        lessons_with_counts.append({
            'lesson': lesson,
            'lesson_count': lesson_count
        })

    return render_template('exam_cards.html', lessons_with_counts=lessons_with_counts,sector_code=sector_code)


@app.route('/lesson_cards', methods=['GET', 'POST'])
def lesson_cards():
    employee_id = session.get('employee_id')
    sector_code = session.get('sector_code')
    if not sector_code or not employee_id:
        return redirect(url_for('access'))
    chapter_id = request.args.get('chapter_id', type=int)
    lessons = Chapter.query.filter_by(id_chapter=chapter_id).all()
    return render_template('lesson_cards.html', lessons=lessons,sector_code=sector_code)


@app.route('/add_lessons', methods=['GET', 'POST'])
def add_lessons():
    title = request.form.get('title') 
    image = request.files.get('image')  
    employee_id = session.get('employee_id')
    print(employee_id)
    print(title)
    
    chapter_id = request.args.get('chapter_id', type=int) 
    print(chapter_id)
   
    if not title or not employee_id or not chapter_id:
        return "Title, Employee ID, and Chapter ID are required", 400

  
    new_lesson = Chapter(
        title=title,
        employee_id=employee_id,
        id_chapter=chapter_id,
        image_path=None  
    )
    db.session.add(new_lesson)
    db.session.commit() 

    if image:
        image_path = f"static/uploads/{new_lesson.id}_L_{image.filename}"
        image.save(image_path)
        
        new_lesson.image_path = image_path
        db.session.commit()

    return redirect(url_for('lesson_cards', chapter_id=chapter_id))


@app.route('/edit_lessons/<int:lesson_id>', methods=['POST'])
def edit_lessons(lesson_id):
    lesson = Chapter.query.get_or_404(lesson_id)
    chapter_id = lesson.id_chapter
    title = request.form.get('title')

    if title:
        lesson.title = title
        db.session.commit()

    return redirect(url_for('lesson_cards', chapter_id=chapter_id))

@app.route('/delete_lessons/<int:lesson_id>', methods=['POST'])
def delete_lessons(lesson_id):
    lesson = Chapter.query.get_or_404(lesson_id)
    chapter_id = lesson.id_chapter
    if lesson.image_path:
        os.remove(lesson.image_path) 
    db.session.delete(lesson)
    db.session.commit()
    return redirect(url_for('lesson_cards', chapter_id=chapter_id))

@app.route('/add_lesson', methods=['GET', 'POST'])
def add_lesson():
    title = request.form.get('title') 
    image = request.files.get('image')  
    employee_id = session.get('employee_id')
    if not title or not employee_id:
        return "Title and Employee ID are required", 400

    new_lesson = category(
        title=title,
        employee_id=employee_id,
        image_path=None  
    )
    db.session.add(new_lesson)
    db.session.commit() 

    if image:
        image_path = f"static/uploads/{new_lesson.id}_L_{image.filename}"
        image.save(image_path)

        new_lesson.image_path = image_path
        db.session.commit()

    return redirect(url_for('exam_cards'))

@app.route('/edit_lesson/<int:lesson_id>', methods=['POST'])
def edit_lesson(lesson_id):
    lesson = category.query.get_or_404(lesson_id)
    title = request.form.get('title')

    if title:
        lesson.title = title
        db.session.commit()

    return redirect(url_for('exam_cards'))


@app.route('/delete_lesson/<int:lesson_id>', methods=['POST'])
def delete_lesson(lesson_id):
    lesson = category.query.get_or_404(lesson_id)
    if lesson.image_path:
        os.remove(lesson.image_path) 
    db.session.delete(lesson)
    db.session.commit()
    return redirect(url_for('exam_cards'))

@app.route('/company_cards', methods=['GET', 'POST'])
def company_cards():
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    sector_code = session.get('sector_code')
    branches = CompanyName.query.filter_by(code=sector_code).all()
    col3 = Administrative_Title.query.filter_by(code=sector_code).all()
    col4 = JobTitle.query.filter_by(code=sector_code).all()
    col5 = specializationEmployee.query.filter_by(code=sector_code).all()
    data = []
    employees = Employee6.query.filter_by(code=sector_code).all()
    department_employees = DepartmentEmployee.query.filter_by(code=sector_code).all()
    department_employees1 = DepartmentEmployee1.query.filter_by(code=sector_code).all()
    for branch in branches:
        branch_name = branch.name
        branch_code = branch.code

        total_employees = Employee6.query.filter_by(branch=branch_name,code=sector_code).count()
        total_managers = Employee6.query.filter_by(branch=branch_name,code=sector_code,is_manager=True).count()
        total_heads = Employee6.query.filter_by(branch=branch_name,code=sector_code,is_head="ttt").count()

        head_details = head.query.filter_by(branch=branch_name,code=sector_code).all()
        manager_details = ManagerEmployee.query.filter_by(branch=branch_name,code=sector_code).all()
        employee_details = Employee6.query.filter_by(branch=branch_name,code=sector_code).all()

      
        col2_counts = {}
        department_values = [d.name for d in department_employees]
        for value in department_values:
            count = Employee6.query.filter_by(col2=value, branch=branch_name).count()
            col2_counts[value] = count
        
        col1_counts = {}
        department_values = [d.name for d in department_employees1]
        for value in department_values:
            count = Employee6.query.filter_by(col1=value, branch=branch_name).count()
            col1_counts[value] = count
        col3_counts = {}
        department_values = [d.name for d in col3]
        for value in department_values:
            count = Employee6.query.filter_by(col3=value, branch=branch_name).count()
            col3_counts[value] = count

        col4_counts = {}
        department_values = [d.name for d in col4]
        for value in department_values:
            count = Employee6.query.filter_by(col4=value, branch=branch_name).count()
            col4_counts[value] = count

        col5_counts = {}
        department_values = [d.name for d in col5]
        for value in department_values:
            count = Employee6.query.filter_by(col5=value, branch=branch_name).count()
            col5_counts[value] = count

        data.append({
            'branch_name': branch_name,
            'total_employees': total_employees,
            'total_managers': total_managers,
            'total_heads': total_heads,
            'head_details': head_details,
            'manager_details': manager_details,
            'employee_details': employee_details,
            'col2_counts': col2_counts, 
            'col1_counts': col1_counts,
            'col3_counts': col3_counts,
            'col4_counts': col4_counts,
            'col5_counts': col5_counts,
        })

    return render_template('company_cards.html',employees=employees, data=data,sector_id=sector_id,sector_code=sector_code)



@app.route('/company_cards_adm', methods=['GET', 'POST'])
def company_cards_adm():
    user_id = session.get('user_id')
    sector_code = session.get('sector_code')
    branches = CompanyName.query.filter_by(code=sector_code).all()
    col3 = Administrative_Title.query.filter_by(code=sector_code,id_admin=user_id).all()
    col4 = JobTitle.query.filter_by(code=sector_code,id_admin=user_id,).all()
    col5 = specializationEmployee.query.filter_by(code=sector_code,id_admin=user_id).all()
    data = []
    employees = Employee6.query.filter_by(code=sector_code,id_admin=user_id).all()
    department_employees = DepartmentEmployee.query.filter_by(code=sector_code,id_admin=user_id,).all()
    department_employees1 = DepartmentEmployee1.query.filter_by(code=sector_code,id_admin=user_id,).all()
    for branch in branches:
        branch_name = branch.name
        branch_code = branch.code

        total_employees = Employee6.query.filter_by(branch=branch_name,id_admin=user_id).count()
        total_managers = Employee6.query.filter_by(branch=branch_name,id_admin=user_id,is_manager=True).count()
        total_heads = Employee6.query.filter_by(branch=branch_name,id_admin=user_id,is_head="ttt").count()

        head_details = head.query.filter_by(branch=branch_name,id_admin=user_id).all()
        manager_details = ManagerEmployee.query.filter_by(branch=branch_name,id_admin=user_id,).all()
        employee_details = Employee6.query.filter_by(branch=branch_name,id_admin=user_id,).all()

      
        col2_counts = {}
        department_values = [d.name for d in department_employees]
        for value in department_values:
            count = Employee6.query.filter_by(col2=value, branch=branch_name,id_admin=user_id,).count()
            col2_counts[value] = count
        
        col1_counts = {}
        department_values = [d.name for d in department_employees1]
        for value in department_values:
            count = Employee6.query.filter_by(col1=value, branch=branch_name,id_admin=user_id,).count()
            col1_counts[value] = count
        col3_counts = {}
        department_values = [d.name for d in col3]
        for value in department_values:
            count = Employee6.query.filter_by(col3=value, branch=branch_name,id_admin=user_id,).count()
            col3_counts[value] = count

        col4_counts = {}
        department_values = [d.name for d in col4]
        for value in department_values:
            count = Employee6.query.filter_by(col4=value, branch=branch_name,id_admin=user_id,).count()
            col4_counts[value] = count

        col5_counts = {}
        department_values = [d.name for d in col5]
        for value in department_values:
            count = Employee6.query.filter_by(col5=value, branch=branch_name,id_admin=user_id,).count()
            col5_counts[value] = count

        data.append({
            'branch_name': branch_name,
            'total_employees': total_employees,
            'total_managers': total_managers,
            'total_heads': total_heads,
            'head_details': head_details,
            'manager_details': manager_details,
            'employee_details': employee_details,
            'col2_counts': col2_counts, 
            'col1_counts': col1_counts,
            'col3_counts': col3_counts,
            'col4_counts': col4_counts,
            'col5_counts': col5_counts,
        })

    return render_template('company_cards_adm.html', data=data,sector_code=sector_code,employees=employees)

ALLOWED_EXTENSIONS_T_F = {'png', 'jpg', 'jpeg', 'gif', 'mp3', 'wav'}

def allowed_file1(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS_T_F

@app.route('/true_false', methods=['GET', 'POST'])
def true_false():
    question_id = request.args.get('question_id', type=int)
    question = Questionfor_t_or_f.query.get_or_404(question_id)

    if request.method == 'POST':
        # تحديث البيانات الأخرى
        question.question_name = request.form['question_name']
        question.difficulty = request.form['difficulty']
        question.importance = request.form['importance']
        question.bloom_taxonomy = request.form['bloom_taxonomy']
        question.grade = request.form['grade']
        question.times = request.form['times']
        correct_answer_value = request.form['correct_answer'] == 'True'
        question.correct_answer = correct_answer_value

       
        # معالجة رفع الصورة
        if 'photo' in request.files:
            photo = request.files['photo']
            if photo and allowed_file1(photo.filename):
                # حذف الصورة القديمة إذا كانت موجودة
                if question.photo_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path))
                    except FileNotFoundError:
                        pass

                # تحديد الاسم الجديد للصورة مع تغيير الامتداد إلى .png
                filename = f"T_F_{question_id}_{secure_filename(photo.filename).rsplit('.', 1)[0]}.png"
                try:
                    photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.photo_path = filename
                except Exception as e:
                    print(f"Error saving photo: {e}")

        # معالجة رفع الصوت
        if 'audio' in request.files:
            audio = request.files['audio']
            if audio and allowed_file1(audio.filename):
                # حذف الصوت القديم إذا كان موجودًا
                if question.audio_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path))
                    except FileNotFoundError:
                        pass

                # تحديد الاسم الجديد للصوت مع الحفاظ على الامتداد الأصلي
                filename = f"T_F_{question_id}_{secure_filename(audio.filename)}"
                try:
                    audio.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.audio_path = filename
                except Exception as e:
                    print(f"Error saving audio: {e}")

        # حفظ التغييرات في قاعدة البيانات
        try:
            db.session.commit()
        except Exception as e:
            print(f"Error saving data: {e}")

    return render_template('true_false.html', question=question)

@app.route('/signup_sector', methods=['GET', 'POST'])
def signup_sector():
    if request.method == 'POST':
        foundation_name = request.form.get('foundation_name')
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')

        if not foundation_name or len(foundation_name) < 3:
            flash("اسم الشركة يجب أن يكون أطول من 3 أحرف", "error")
            return render_template('signup_sector.html')
        
        if not username or len(username) < 3:
            flash("اسم المستخدم يجب أن يكون أطول من 3 أحرف", "error")
            return render_template('signup_sector.html')

        email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
        if not email or not re.match(email_regex, email):
            flash("صيغة البريد الإلكتروني غير صحيحة", "error")
            return render_template('signup_sector.html')

        if not email or Sector.query.filter_by(email=email).first():
            flash("الإيميل مستخدم بالفعل أو غير صالح", "error")
            return render_template('signup_sector.html')
        
        if not password or len(password) < 6:
            flash("كلمة المرور يجب أن تكون أطول من 6 أحرف", "error")
            return render_template('signup_sector.html')

      
        code = None
        while not code or Sector.query.filter_by(code=code).first():
            code = str(random.randint(10**7, 10**8 - 1))
        
        
        new_sector = Sector(
            foundation_name=foundation_name,
            username=username,
            email=email,
            password=password,
            code=code
        )

        db.session.add(new_sector)
        db.session.commit()
        new_notification = Notification_owner(
        subject=f"Hello {username}",
        message = "Welcome to Noor Academy for Sustainable Development, where we focus on nurturing human potential sustainably. We wish you a successful and continuous learning journey filled with growth .",
        company_code=code,
        timestamp=datetime.now()
        )
        db.session.add(new_notification)
        db.session.commit()


        session['sector_id'] = new_sector.id
        session['sector_code'] = new_sector.code

        return redirect(url_for('owner'))

    return render_template('signup_sector.html')
@app.route('/signup_personal', methods=['GET', 'POST'])
def signup_personal():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        account_type = request.form.get('account_type')

        if username and email and password and account_type:
            new_personal = Personal(
                username=username,
                email=email,
                password=password,
                account_type=account_type
            )
            db.session.add(new_personal)
            db.session.commit()

            return redirect(url_for('success', username=username))

    return render_template('signup_personal.html')



@app.route('/', methods=['GET', 'POST'])
def index():
    return render_template('index.html')    

@app.route('/login', methods=['GET', 'POST'])
def login():
    return render_template('login.html')    

@app.route('/sector_login', methods=['GET', 'POST'])
def sector_login():
    return render_template('sector_login.html')

@app.route('/login_sector', methods=['GET', 'POST'])
def login_sector():
    if request.method == 'POST':
        email = request.form.get('email')
        print(email)
        password = request.form.get('password')

        sector = Sector.query.filter_by(email=email).first()
        if not sector:
            flash("الإيميل غير مسجل في النظام", "error")
            return render_template('login_sector.html')

        if sector.password != password:
            flash("كلمة المرور غير صحيحة", "error")
            return render_template('login_sector.html')

        session['sector_id'] = sector.id
        session['sector_code'] = sector.code

        return redirect(url_for('owner'))

    return render_template('login_sector.html')


@app.route('/signup_eml', methods=['GET', 'POST'])
def signup_eml():
    if request.method == 'POST':
        sector_code = request.args.get('sector_code') 
        id_admin = request.args.get('id') 
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')

        if not username or len(username) < 3:
            flash("اسم المستخدم قصير جدًا. يجب أن يكون 3 أحرف على الأقل.", "error")
        elif not email or not re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', email):
            flash("صيغة البريد الإلكتروني غير صحيحة.", "error")
        elif Employee6.query.filter_by(email=email).first():
            flash("البريد الإلكتروني مستخدم بالفعل.", "error")
        elif not password or len(password) < 6:
            flash("كلمة المرور قصيرة جدًا. يجب أن تكون على الأقل 6 أحرف.", "error")
        else:
            try:
                new_employee = Employee6(code=sector_code, username=username, email=email,id_admin=id_admin, password=password)
                db.session.add(new_employee)
                db.session.commit()
                new_notification = Notification_admin(
                subject=f"New Employee Added",
                message=f"A new employee {username}, has been added to Noor Academy for Sustainable Development.",
                id_user=id_admin,
                company_code=sector_code,
                timestamp=datetime.now()
                )
                db.session.add(new_notification)
                db.session.commit()

                session['employee_id'] = new_employee.id
                session['employee_code'] = new_employee.code

                return redirect('/employee_profile')
            except Exception as e:
                db.session.rollback()
                flash("حدث خطأ أثناء إنشاء الحساب. يرجى المحاولة مرة أخرى.", "error")

    return render_template('signup_eml.html')
@app.route('/signup_organizer', methods=['GET', 'POST'])
def signup_organizer():
    code = request.args.get('code') 
    admin_data = ColOrganizer.query.filter_by(code=code).first()

    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        col1 = request.form.get('c1')
        col2 = request.form.get('c2')
        col3 = request.form.get('c3')
        col4 = request.form.get('c4')
        col5 = request.form.get('c5')
        password = request.form['password']
        sector_code = request.args.get('code') 
        
        email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
        if not email or not re.match(email_regex, email):
            flash("صيغة البريد الإلكتروني غير صحيحة", "error")
            return render_template('signup_organizer.html',
                                   c1=admin_data.c1,
                                   c2=admin_data.c2,
                                   c3=admin_data.c3,
                                   c4=admin_data.c4,
                                   c5=admin_data.c5)

        # تحقق من إذا كان البريد الإلكتروني مستخدمًا بالفعل
        existing_organizer = organizer2.query.filter_by(email=email).first()
        if existing_organizer:
            flash("البريد الإلكتروني مستخدم بالفعل", "error")
            return render_template('signup_organizer.html',
                                   c1=admin_data.c1,
                                   c2=admin_data.c2,
                                   c3=admin_data.c3,
                                   c4=admin_data.c4,
                                   c5=admin_data.c5)

        # تحقق من اسم المستخدم
        if len(username) < 3:
            flash("اسم المستخدم قصير جدًا. يجب أن يكون 3 أحرف على الأقل.", "error")
            return render_template('signup_organizer.html',
                                   c1=admin_data.c1,
                                   c2=admin_data.c2,
                                   c3=admin_data.c3,
                                   c4=admin_data.c4,
                                   c5=admin_data.c5)

        # تحقق من كلمة المرور
        if len(password) < 6:
            flash("كلمة المرور قصيرة جدًا. يجب أن تكون على الأقل 6 أحرف.", "error")
            return render_template('signup_organizer.html',
                                   c1=admin_data.c1,
                                   c2=admin_data.c2,
                                   c3=admin_data.c3,
                                   c4=admin_data.c4,
                                   c5=admin_data.c5)

        # إنشاء منظم جديد
        new_organizer = organizer2(
            name=username,
            email=email,
            col1=col1,
            col2=col2,
            col3=col3,
            col4=col4,
            col5=col5,
            password=password,
            code=sector_code
        )
        
        db.session.add(new_organizer)
        db.session.commit()


        new_notification = Notification_owner(
            subject="New Organizer Added",
            message=f"Organizer {username} has been added successfully.",
            company_code=sector_code,
            timestamp=datetime.now()
        )
        db.session.add(new_notification)
        db.session.commit()
        

        session['organizer_id'] = new_organizer.id
        session['sector_code'] = sector_code

        return redirect(url_for('organizer'))


    return render_template('signup_organizer.html',
                           c1=admin_data.c1,
                           c2=admin_data.c2,
                           c3=admin_data.c3,
                           c4=admin_data.c4,
                           c5=admin_data.c5)

@app.route('/notification', methods=['GET'])
def notification():
    
    company_code = session.get('sector_code')
    sector_code = session.get('sector_code')
    notifications = Notification_owner.query.filter_by(company_code=company_code).order_by(desc(Notification_owner.id)).all()


    for notification in notifications:
        if not notification.viewed: 
            notification.viewed = True

    db.session.commit()

    return render_template('notification.html', notifications=notifications,sector_code=sector_code)

@app.route('/notification_organizer', methods=['GET'])
def notification_organizer():
    
    company_code = session.get('sector_code')
    sector_code = session.get('sector_code')
    notifications = Notification_organizer.query.filter_by(company_code=company_code).order_by(desc(Notification_organizer.id)).all()

    for notification in notifications:
        if not notification.viewed: 
            notification.viewed = True

    db.session.commit()

    return render_template('notification_organizer.html', notifications=notifications,sector_code=sector_code)

@app.route('/notification_admin', methods=['GET'])
def notification_admin():
    user_id = session.get('user_id')
    company_code = session.get('sector_code')
    sector_code = session.get('sector_code')
    notifications = Notification_admin.query.filter_by(company_code=company_code, id_user=user_id).order_by(desc(Notification_admin.id)).all()

    for notification in notifications:
        if not notification.viewed: 
            notification.viewed = True

    db.session.commit()

    return render_template('notification_admin.html', notifications=notifications,sector_code=sector_code)

@app.route('/sector_register', methods=['GET', 'POST'])
def sector_register():

    return render_template('sector_register.html')



@app.route('/signup_administrator', methods=['GET', 'POST'])
def signup_administrator():
    sector_code = request.args.get('code')

    if request.method == 'POST':
        name = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        col1 = request.form.get('c1')
        col2 = request.form.get('c2')
        col3 = request.form.get('c3')
        col4 = request.form.get('c4')
        col5 = request.form.get('c5')
        code = request.args.get('code')

        if not name or len(name) < 3:
            flash("اسم المستخدم قصير جدًا. يجب أن يكون 3 أحرف على الأقل.", "error")
        elif not email or not re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', email):
            flash("صيغة البريد الإلكتروني غير صحيحة.", "error")
        elif  administrator2.query.filter_by(email=email).first():
            flash("البريد الإلكتروني مستخدم بالفعل.", "error")
        elif not password or len(password) < 6:
            flash("كلمة المرور قصيرة جدًا. يجب أن تكون على الأقل 6 أحرف.", "error")
        else:
         
            new_admin =  administrator2(
                name=name,
                email=email,
                col1=col1,
                col2=col2,
                col3=col3,
                col4=col4,
                col5=col5,
                code=code,
                password=password
            )
            db.session.add(new_admin)
            db.session.commit()
            new_notification = Notification_owner(
            subject="New Administrator Added",
            message=f"Administrator {name} has been added successfully.",
            company_code=sector_code,
            timestamp=datetime.now()
            
            )
            db.session.add(new_notification)
            db.session.commit()

            new_notification = Notification_organizer(
            subject="New Administrator Added",
            message=f"Administrator {name} has been added successfully.",
            company_code=sector_code
            ,
            timestamp=datetime.now()
            )
            db.session.add(new_notification)
            db.session.commit()

            new_notification = Notification_admin(
                subject=f"Hello {name}",
                message = f"Welcome to Noor Academy for Sustainable Development, where we focus on nurturing human potential sustainably. We wish you a successful and continuous learning journey filled with growth.",
                id_user=new_admin.id,
                company_code=sector_code,
                timestamp=datetime.now()
            )
            db.session.add(new_notification)
            db.session.commit()


            session['user_id'] = new_admin.id
            session['sector_code'] = code

             
            return redirect('/administrator')

    admin_data = ColAdministrator.query.filter_by(code=sector_code).first()
    if not admin_data:
        return "Invalid code", 404

    return render_template(
        'signup_administrator.html',
        c1=admin_data.c1,
        c2=admin_data.c2,
        c3=admin_data.c3,
        c4=admin_data.c4,
        c5=admin_data.c5
    )

@app.route('/employee_profile', methods=['GET', 'POST'])
def employee_profile():
    employee_id = session.get('employee_id')
    sector_code = session.get('sector_code')
    if not sector_code or not employee_id:
        return redirect(url_for('access'))
    employee_data = Employee6.query.filter_by(id=employee_id).first()
    col1 = DepartmentEmployee1.query.filter_by(code=sector_code,id_admin=employee_data.id_admin).all()
    col2 = DepartmentEmployee.query.filter_by(code=sector_code,id_admin=employee_data.id_admin).all()
    col3 = Administrative_Title.query.filter_by(code=sector_code,id_admin=employee_data.id_admin).all()
    col4 = JobTitle.query.filter_by(code=sector_code,id_admin=employee_data.id_admin).all()
    col5 = specializationEmployee.query.filter_by(code=sector_code,id_admin=employee_data.id_admin).all()
    employee_name = ColEmployee.query.filter_by(code=sector_code,id_admin=employee_data.id_admin).first()
    
    print(employee_data.id_admin)
    
    CompanyNames = CompanyName.query.filter_by(code=sector_code).all()
    if not employee_id or not sector_code:
        return redirect(url_for('access')) 
    if not employee_data:
        return redirect('/')

    if request.method == 'POST':
        email = request.form.get('email')
        phone = request.form.get('phone')
        email_pattern = r"(^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$)"
        if not re.match(email_pattern, email):
            flash("البريد الإلكتروني غير صالح")
            return redirect(request.url) 
        phone_pattern = r"^\d+$"
        if not re.match(phone_pattern, phone):
            flash("رقم الهاتف غير صالح. يجب أن يحتوي على أرقام فقط")
            return redirect(request.url)
        employee_data.username = request.form.get('username')
        employee_data.email = request.form.get('email')
        birthdate_str = request.form.get('birthdate')
        employee_data.birthdate = datetime.strptime(birthdate_str, '%Y-%m-%d').date()
        employee_data.major = request.form.get('major')
        employee_data.phone = request.form.get('phone')
        employee_data.country = request.form.get('country')
        employee_data.province = request.form.get('province')

        employee_data.col1 = request.form.get('col1')
        employee_data.col2 = request.form.get('col2')
        employee_data.col3 = request.form.get('col3')
        employee_data.col4 = request.form.get('col4')
        employee_data.col5 = request.form.get('col5')
        employee_data.branch = request.form.get('branch')
       
        db.session.commit()

    return render_template('employee_profile.html', sector_code=sector_code,
                           employee_data=employee_data,
                           col1=col1, col2=col2, 
                           col3=col3, col4=col4, 
                           col5=col5, employee_name=employee_name, 
                           CompanyNames=CompanyNames)



@app.route('/Studentprofile', methods=['GET', 'POST'])
def Studentprofile():
    student_id=session.get('student_id')
    sector_code = session.get('sector_code')
    if not sector_code or not student_id:
        return redirect(url_for('access'))
    student_data = SectorStudent8.query.filter_by(id=student_id).first() 
    student_name = ColStudent.query.filter_by(code=sector_code,id_admin=student_data.id_admin).first()
    print(student_data.id_admin)
    if not student_id or not sector_code:
        return redirect(url_for('access')) 
    col1 = yearStudent.query.filter_by(code=sector_code,id_admin=student_data.id_admin).all()
    col2 = DepartmentStudent.query.filter_by(code=sector_code,id_admin=student_data.id_admin).all()
    col3 = ClassStudent.query.filter_by(code=sector_code,id_admin=student_data.id_admin).all()
    col4 = GradeStudent.query.filter_by(code=sector_code,id_admin=student_data.id_admin).all()
    CompanyNames = CompanyName.query.filter_by(code=sector_code).all()
    col5 = specializationStudent.query.filter_by(code=sector_code,id_admin=student_data.id_admin).all()

    if request.method == 'POST':
        student_data.username = request.form.get('username')
        student_data.email = request.form.get('email')
        email = request.form.get('email')
        phone = request.form.get('phone')
        email_pattern = r"(^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$)"
        if not re.match(email_pattern, email):
            flash("البريد الإلكتروني غير صالح")
            return redirect(request.url) 
        phone_pattern = r"^\d+$"
        if not re.match(phone_pattern, phone):
            flash("رقم الهاتف غير صالح. يجب أن يحتوي على أرقام فقط")
            return redirect(request.url)
        birthdate_str = request.form.get('birthdate')
        student_data.birthdate = datetime.strptime(birthdate_str, '%Y-%m-%d').date()
        student_data.stage = request.form.get('stage')
        student_data.university = request.form.get('university')
        student_data.major = request.form.get('major')
        student_data.phone = request.form.get('phone')
        student_data.country = request.form.get('country')
        student_data.province = request.form.get('province')
        student_data.col1 = request.form.get('col1')
        student_data.col2 = request.form.get('col2')
        student_data.col3 = request.form.get('col3')
        student_data.col4 = request.form.get('col4')
        student_data.col5 = request.form.get('col5')
        db.session.commit()
    return render_template('Studentprofile.html',sector_code=sector_code,student_data=student_data,student_name=student_name,
    col1=col1,col2=col2,col3=col3,col4=col4,col5=col5,CompanyNames=CompanyNames)

@app.route('/delete_job_title/<int:job_id>', methods=['GET'])
def delete_job_title(job_id):
    job = JobTitle.query.get(job_id)
    if job:
        db.session.delete(job)
        db.session.commit()
    return redirect(url_for('employee'))

@app.route('/delete_specialization/<int:specialization_id>', methods=['GET'])
def delete_specialization(specialization_id):
    specialization = specializationEmployee.query.get(specialization_id)
    if specialization:
        db.session.delete(specialization)
        db.session.commit()
    return redirect(url_for('employee'))

@app.route('/signup_student', methods=['GET', 'POST'])
def signup_student():
    
    sector_code = request.args.get('sector_code') 
    id_admin = request.args.get('id')  
    print(sector_code)
    print(id_admin)
  
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')

        # التحقق من المدخلات
        if not username or len(username) < 3:
            flash("اسم المستخدم قصير جدًا. يجب أن يكون 3 أحرف على الأقل.", "error")
        elif not email or not re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', email):
            flash("صيغة البريد الإلكتروني غير صحيحة.", "error")
        elif SectorStudent8.query.filter_by(email=email).first():
            flash("البريد الإلكتروني مستخدم بالفعل.", "error")
        elif not password or len(password) < 6:
            flash("كلمة المرور قصيرة جدًا. يجب أن تكون على الأقل 6 أحرف.", "error")
        else:
           
            new_sector = SectorStudent8(
                username=username,
                email=email,
                password=password,
                code=sector_code,
                id_admin=id_admin,
                accepted=False
            )
            db.session.add(new_sector)
            db.session.commit()
            new_notification = Notification_admin(
                subject=f"New Student Added",
                message=f"A new student {username}, has been added to Noor Academy for Sustainable Development. The student is currently awaiting your approval.",
                id_user=id_admin,
                company_code=sector_code,
                timestamp=datetime.now()
            )
            db.session.add(new_notification)
            db.session.commit()

            session['student_id'] = new_sector.id
            session['sector_code'] = new_sector.code

            return redirect(url_for('Studentprofile'))

    departments = DepartmentStudent.query.filter_by(code=sector_code).all()
    ClassStudents = ClassStudent.query.filter_by(code=sector_code).all()

    return render_template('signup_student.html', sector_code=sector_code,id=id_admin, departments=departments, ClassStudents=ClassStudents)

@app.route('/validate_code', methods=['POST'])
def validate_code():
    user_code = request.form.get('codeInput')
    for i, char in enumerate(user_code):
            if char.isalpha():
                sector_code = user_code[:i]  
                person_code = user_code[i + 1:]  
                user_type = char.upper()  
                break

    sector = Sector.query.filter_by(code=sector_code).first()  

    if sector:
        if user_type == 'S': 
            return redirect(url_for('signup_student', sector_code=sector_code, id=person_code))
        elif user_type == 'E':  
            return redirect(url_for('signup_eml', sector_code=sector_code, id=person_code))
        else:
            if user_type == 'A':
                return redirect(url_for('signup_administrator', code=sector_code))
            elif user_type == 'M':  
                return redirect(url_for('signup_organizer', code=sector_code))
            else:
                print("Invalid person code.")
                return redirect(url_for('sector_register'))
    else:
        print("Sector code is not valid. Please try again.", "error")
        return redirect(url_for('sector_register'))


@app.route('/validate_code1', methods=['POST'])
def validate_code1():
    
    user_code = request.form.get('codeInput')

    if user_code == 'O':
        return redirect(url_for('login_organizer')) 
    elif user_code == 'A':
        return redirect(url_for('login_administrator')) 
    elif user_code == 'S':
        return redirect(url_for('login_Student'))  
    elif user_code == 'E':
        return redirect(url_for('login_employee'))  
    else:
        flash("الرمز غير صالح", "error")



@app.route('/login_organizer', methods=['GET', 'POST'])
def login_organizer():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        organizer = organizer2.query.filter_by(email=email).first()
        if not organizer:
            flash("البريد الإلكتروني غير مسجل في النظام", "error")
            return render_template('login_organizer.html')

        if organizer.password != password:
            flash("كلمة المرور غير صحيحة", "error")
            return render_template('login_organizer.html')

        session['organizer_id'] = organizer.id
        session['sector_code'] = organizer.code

        return redirect(url_for('organizer'))

    return render_template('login_organizer.html')
    
    
@app.route('/login_administrator', methods=['GET', 'POST'])
def login_administrator():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        admin =  administrator2.query.filter_by(email=email).first()
        if not admin:
            flash("البريد الإلكتروني غير مسجل في النظام", "error")
            return render_template('login_administrator.html')

        if admin.password != password:
            flash("كلمة المرور غير صحيحة", "error")
            return render_template('login_administrator.html')

        session['user_id'] = admin.id
        session['sector_code'] = admin.code

        return redirect('/administrator')

    return render_template('login_administrator.html')


@app.route('/login_Student', methods=['GET', 'POST'])
def login_Student():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        student = SectorStudent8.query.filter_by(email=email).first()
        if not student:
            flash("البريد الإلكتروني غير مسجل في النظام.", "error")
            return render_template('login_Student.html')

        if student.password != password:
            flash("كلمة المرور غير صحيحة.", "error")
            return render_template('login_Student.html')


        session['student_id'] = student.id
        session['sector_code'] = student.code

        return redirect(url_for('Studentprofile'))

    return render_template('login_Student.html')

@app.route('/login_employee', methods=['GET', 'POST'])
def login_employee():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        employee = Employee6.query.filter_by(email=email).first()

        if not employee:
            flash("البريد الإلكتروني غير مسجل في النظام.", "error")
            return render_template('login_employee.html')

        if employee.password != password:
            flash("كلمة المرور غير صحيحة.", "error")
            return render_template('login_employee.html')

        session['employee_id'] = employee.id
        session['employee_code'] = employee.code

        
        return redirect('/employee_profile')

    return render_template('login_employee.html')


@app.route('/organizer', methods=['GET', 'POST'])
def organizer():
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    unread_notifications = Notification_organizer.query.filter_by(company_code=sector_code, viewed=False).all()
    if unread_notifications:
        notification_dot_display = True
    else:
        notification_dot_display = False
    if not sector_code or not (organizer_id):
        return redirect(url_for('access'))
    admin_data = ColOrganizer.query.filter_by(code=sector_code).first()
    organizer_data = organizer2.query.filter_by(id=organizer_id).first()
    administrators =  administrator2.query.filter_by(code=sector_code).all()
    sector_id = session.get('sector_id')
    sector_code = session.get('sector_code')
    column_names = ColAdministrator.query.filter_by(code=sector_code).first()
    company_names = CompanyName.query.filter_by(code=sector_code).all()
    print(company_names)
    col_data = ColAdministrator.query.filter_by(code=sector_code).first()
    if request.method == 'POST' and request.form.get('company_name')!=None:
        company_name = request.form.get('company_name')

        if company_name and sector_code:
            existing_company = CompanyName.query.filter_by(name=company_name, code=sector_code).first()
            if existing_company:
                return render_template('owner.html',notification_dot_display=notification_dot_display,sector_code=sector_code, company_names=company_names, error="اسم الشركة وكودها موجودان بالفعل")
            else:
                try:
                    new_company = CompanyName(name=company_name, code=sector_code)
                    db.session.add(new_company)
                    db.session.commit()
                    return redirect(url_for('organizer')) 
                except exc.IntegrityError as e:
                    db.session.rollback()
                    return render_template('organizer.html',notification_dot_display=notification_dot_display,sector_code=sector_code, company_names=company_names, error=f"خطأ في قاعدة البيانات: {e}")
                except Exception as e:
                    db.session.rollback()
                    return render_template('organizer.html',notification_dot_display=notification_dot_display,sector_code=sector_code, company_names=company_names, error=f"حدث خطأ: {e}")
    if 'excel_file' in request.files:
        excel_file = request.files['excel_file']
        if excel_file.filename != '':
            filename = secure_filename(excel_file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            excel_file.save(filepath)

            try:
                workbook = load_workbook(filepath)
                sheet = workbook.active  

                for row in sheet.iter_rows(min_row=2):
                    name = row[0].value
                    print(name) 
                    print(sector_code)

                    existing_company = CompanyName.query.filter_by(name=name, code=sector_code).first()

                    if existing_company:
                        print(f"Duplicate entry detected: {name} with code {sector_code}")
                        continue 

                    new_company = CompanyName(name=name, code=sector_code)
                    db.session.add(new_company)

                db.session.commit()
            except exc.IntegrityError as e:
                db.session.rollback()
                print(f"Database integrity error: {e}")
            except Exception as e:
                db.session.rollback()
                print(f"An error occurred: {e}")
                print(f"Error processing Excel file: {e}")
                return render_template('organizer.html',unread_notifications=unread_notifications,organizer=organizer_data,administrators=administrators, columns=column_names,sector_code=sector_code,col_data=col_data,  error=f"Error processing Excel file: {e}")
            finally:
                os.remove(filepath)
    if request.method == 'POST' and  request.form.get('name'):


            new_name = request.form['name']
            new_email = request.form['email']
            organizer_data.col1 = request.form.get('c1', organizer_data.col1)
            organizer_data.col2 = request.form.get('c2', organizer_data.col2)
            organizer_data.col3 = request.form.get('c3', organizer_data.col3)
            organizer_data.col4 = request.form.get('c4', organizer_data.col4)
            organizer_data.col5 = request.form.get('c5', organizer_data.col5)

            organizer_data.name = new_name
            organizer_data.email = new_email
            
            db.session.commit() 
    if not col_data:
        col_data = ColAdministrator(code=sector_code, c1='', c2='', c3='', c4='', c5='')
        db.session.add(col_data)
        db.session.commit()
    if request.method == 'POST':
        col_data.c1 = request.form.get('column1', col_data.c1)
        col_data.c2 = request.form.get('column2', col_data.c2)
        col_data.c3 = request.form.get('column3', col_data.c3)
        col_data.c4 = request.form.get('column4', col_data.c4)
        col_data.c5 = request.form.get('column5', col_data.c5)
        db.session.commit()
        return redirect(url_for('organizer'))

     
    
    return render_template('organizer.html',notification_dot_display=notification_dot_display,organizer_id=organizer_id,organizer=organizer_data, columns=column_names,sector_code=sector_code,col_data=col_data,administrators=administrators,  company_names=company_names
    ,
    c1=admin_data.c1,
    c2=admin_data.c2,
    c3=admin_data.c3,
    c4=admin_data.c4,
    c5=admin_data.c5)    
    

@app.route('/Education_entity', methods=['GET', 'POST'])
def Education_entityr():
    return render_template('Education_entity.html')
@app.route('/delete_student/<int:student_id>', methods=['GET','POST'])
def delete_student(student_id):
    student = SectorStudent8.query.get_or_404(student_id)
    db.session.delete(student)
    db.session.commit()
    return redirect(url_for('administrator')) 
@app.route('/administrator', methods=['GET', 'POST'])
def administrator():

    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    unaccepted_students = SectorStudent8.query.filter_by(accepted=False, code=sector_code,id_admin=user_id).all()
    accepted_students = SectorStudent8.query.filter_by(accepted=True, code=sector_code,id_admin=user_id).all()
    if not sector_code or not (user_id):
        return redirect(url_for('access'))
    col_name = ColAdministrator.query.filter_by(code=sector_code).first()
   
    col_data =  administrator2.query.filter_by(id=user_id).first()
    unread_notifications = Notification_admin.query.filter_by(company_code=sector_code, viewed=False,id_user=user_id).all()
    print(unread_notifications)
    if unread_notifications:
        notification_dot_display = True
    else:
        notification_dot_display = False
 
    if request.method == 'POST':
     
        col_data.name = request.form.get('name', col_data.name)
        col_data.email = request.form.get('email', col_data.email)
        col_data.col1 = request.form.get('c1', col_data.col1)
        col_data.col2 = request.form.get('c2', col_data.col2)
        col_data.col3 = request.form.get('c3', col_data.col3)
        col_data.col4 = request.form.get('c4', col_data.col4)
        col_data.col5 = request.form.get('c5', col_data.col5)

        db.session.commit()


        return redirect(url_for('administrator',notification_dot_display=notification_dot_display))

    return render_template('administrator.html',notification_dot_display=notification_dot_display, sector_code=sector_code, col_data=col_data,col_name=col_name,unaccepted_students=unaccepted_students,
        accepted_students=accepted_students,user_id=user_id
    )

@app.route('/accept_student/<int:student_id>', methods=['GET','POST'])
def accept_student(student_id):
    student = SectorStudent8.query.get(student_id)
    if student:
        student.accepted = True
        db.session.commit()
    return redirect(url_for('administrator'))

@app.route('/delete_sector/<int:sector_id>', methods=['GET'])
def delete_sector(sector_id):
    sector = organizer2.query.get_or_404(sector_id)
    db.session.delete(sector)
    db.session.commit()
    return redirect(url_for('show_org'))

@app.route('/delete_company/<int:company_id>')
def delete_company(company_id):

    company = CompanyName.query.get_or_404(company_id)

    db.session.delete(company)
    db.session.commit()

    return redirect(url_for('organizer'))

@app.route('/delete_year/<int:company_id>')
def delete_year(company_id):

    year_Student = yearStudent.query.get_or_404(company_id)
    
    db.session.delete(year_Student)
    db.session.commit()

    return redirect(url_for('student'))

@app.route('/delete_grades/<int:grade_id>')
def delete_grades(grade_id):

    year_Student = GradeStudent.query.get_or_404(grade_id)
    
    db.session.delete(year_Student)
    db.session.commit()

    return redirect(url_for('student'))

@app.route('/delete_specializations/<int:company_id>')
def delete_specializations(company_id):
   
    year_Student = specializationStudent.query.get_or_404(company_id)
    
    db.session.delete(year_Student)
    db.session.commit()

    return redirect(url_for('student'))

@app.route('/delete_class/<int:class_id>')
def delete_class(class_id):
    year_Student = ClassStudent.query.get_or_404(class_id)
    
    db.session.delete(year_Student)
    db.session.commit()

    return redirect(url_for('student'))


@app.route('/delete-administrative-title/<int:title_id>', methods=['GET', 'POST'])
def delete_administrative_title(title_id):
    title_to_delete = Administrative_Title.query.get_or_404(title_id)

    try:
        db.session.delete(title_to_delete)
        db.session.commit()
    except Exception as e:
        db.session.rollback()

    return redirect(url_for('employee'))

@app.route('/employee', methods=['GET', 'POST'])
def employee():
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    sector_code_display = f"{sector_code}E{user_id}"  
    if not sector_code or not (user_id):
        return redirect(url_for('access'))
    
    student_data = ColEmployee.query.filter_by(code=sector_code,id_admin=user_id).first()
    if not student_data:
        student_data = ColEmployee(code=sector_code,id_admin=user_id, c1="", c2="", c3="", c4="", c5="")
        db.session.add(student_data)
        db.session.commit()
    
    if request.method == 'POST':
        new_c1_value = request.form.get('col1')
        new_c2_value = request.form.get('col2')
        new_c3_value = request.form.get('col3')
        new_c4_value = request.form.get('col4')
        new_c5_value = request.form.get('col5')

        if new_c1_value is not None:
            student_data.c1 = new_c1_value
            db.session.commit()
        
        
        if new_c2_value is not None:
            student_data.c2 = new_c2_value
            db.session.commit()

        if new_c3_value is not None:
            student_data.c3 = new_c3_value
            db.session.commit()

        if new_c4_value is not None:
            student_data.c4 = new_c4_value
            db.session.commit()
        if new_c5_value is not None:
            student_data.c5 = new_c5_value
            db.session.commit()

    if request.method == 'POST':
        job_title = request.form.get('job_title')
        if job_title:
            try:
                existing_job = JobTitle.query.filter_by(name=job_title, code=sector_code , id_admin=user_id).first()
                if not existing_job:
                    new_job = JobTitle(name=job_title, code=sector_code, id_admin=user_id)
                    db.session.add(new_job)
                    db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        if 'job_excel_file' in request.files:
            job_excel_file = request.files['job_excel_file']
            if job_excel_file.filename != '':
                filename = secure_filename(job_excel_file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                job_excel_file.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=2):
                        name = row[0].value
                        if not JobTitle.query.filter_by(name=name, code=sector_code, id_admin=user_id).first():
                            new_job = JobTitle(name=name, code=sector_code, id_admin=user_id)
                            db.session.add(new_job)
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    print(f"Error: {e}")
                finally:
                    os.remove(filepath)

        specialization_name = request.form.get('specialization_name')
        if specialization_name:
            try:
                existing_specialization = specializationEmployee.query.filter_by(name=specialization_name, code=sector_code, id_admin=user_id).first()
                if not existing_specialization:
                    new_specialization = specializationEmployee(name=specialization_name, code=sector_code, id_admin=user_id)
                    db.session.add(new_specialization)
                    db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        if 'specialization_excel_file' in request.files:
            specialization_excel_file = request.files['specialization_excel_file']
            if specialization_excel_file.filename != '':
                filename = secure_filename(specialization_excel_file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                specialization_excel_file.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=2):
                        name = row[0].value
                        if not specializationEmployee.query.filter_by(name=name, code=sector_code, id_admin=user_id).first():
                            new_specialization = specializationEmployee(name=name, code=sector_code, id_admin=user_id)
                            db.session.add(new_specialization)
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    print(f"Error: {e}")
                finally:
                    os.remove(filepath)

        administrative_title_name = request.form.get('administrative_title_name')
        if administrative_title_name:
            try:
                existing_specialization = Administrative_Title.query.filter_by(name=administrative_title_name, code=sector_code, id_admin=user_id).first()
                if not existing_specialization:
                    new_specialization = Administrative_Title(name=administrative_title_name, code=sector_code, id_admin=user_id)
                    db.session.add(new_specialization)
                    db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        if 'administrative_titles_file' in request.files:
            administrative_titles_file = request.files['administrative_titles_file']
            if administrative_titles_file.filename != '':
                filename = secure_filename(administrative_titles_file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                administrative_titles_file.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=2):
                        name = row[0].value
                        if not Administrative_Title.query.filter_by(name=name, code=sector_code, id_admin=user_id).first():
                            new_specialization = Administrative_Title(name=name, code=sector_code, id_admin=user_id)
                            db.session.add(new_specialization)
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    print(f"Error: {e}")
                finally:
                    os.remove(filepath)
            return redirect(url_for('employee'))

        department_name = request.form.get('department_name')
        if department_name:
            try:
                existing_specialization = DepartmentEmployee.query.filter_by(name=department_name, code=sector_code, id_admin=user_id).first()
                if not existing_specialization:
                    new_specialization = DepartmentEmployee(name=department_name, code=sector_code, id_admin=user_id)
                    db.session.add(new_specialization)
                    db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        if 'departments_file' in request.files:
            departments_file = request.files['departments_file']
            if departments_file.filename != '':
                filename = secure_filename(departments_file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                departments_file.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=2):
                        name = row[0].value
                        if not DepartmentEmployee.query.filter_by(name=name, code=sector_code, id_admin=user_id).first():
                            new_specialization = DepartmentEmployee(name=name, code=sector_code, id_admin=user_id)
                            db.session.add(new_specialization)
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    print(f"Error: {e}")
                finally:
                    os.remove(filepath)
            return redirect(url_for('employee'))

        department_name = request.form.get('department_name1')
        if department_name:
            try:
                existing_specialization = DepartmentEmployee1.query.filter_by(name=department_name, code=sector_code, id_admin=user_id).first()
                if not existing_specialization:
                    new_specialization = DepartmentEmployee1(name=department_name, code=sector_code, id_admin=user_id)
                    db.session.add(new_specialization)
                    db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        if 'departments_file1' in request.files:
            departments_file = request.files['departments_file1']
            if departments_file.filename != '':
                filename = secure_filename(departments_file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                departments_file.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=2):
                        name = row[0].value
                        if not DepartmentEmployee1.query.filter_by(name=name, code=sector_code, id_admin=user_id).first():
                            new_specialization = DepartmentEmployee1(name=name, code=sector_code, id_admin=user_id)
                            db.session.add(new_specialization)
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    print(f"Error: {e}")
                finally:
                    os.remove(filepath)
            return redirect(url_for('employee'))

    DepartmentEmployees1 = DepartmentEmployee1.query.filter_by(code=sector_code,id_admin=user_id).all()
    DepartmentEmployees = DepartmentEmployee.query.filter_by(code=sector_code,id_admin=user_id).all()
    Administrative_Titles = Administrative_Title.query.filter_by(code=sector_code,id_admin=user_id).all()
    job_titles = JobTitle.query.filter_by(code=sector_code,id_admin=user_id).all()
    specializations_list = specializationEmployee.query.filter_by(code=sector_code,id_admin=user_id).all()

    return render_template('employee.html',sector_code_display=sector_code_display,student_data=student_data, job_titles=job_titles,sector_code=sector_code, specializations_list=specializations_list,Administrative_Titles=Administrative_Titles,DepartmentEmployees=DepartmentEmployees
    ,DepartmentEmployees1=DepartmentEmployees1)

@app.route('/student', methods=['GET', 'POST'])
def student():
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    if not sector_code or not (user_id):
        return redirect(url_for('access'))
    
    student_data = ColStudent.query.filter_by(code=sector_code,id_admin=user_id).first()

    if not student_data:
        student_data = ColStudent(code=sector_code,id_admin=user_id, c1="", c2="", c3="", c4="", c5="")
        db.session.add(student_data)
        db.session.commit()
    sector_code_display = f"{sector_code}S{user_id}"  
    if request.method == 'POST':
        # تحديث القيم
        updated = False
        new_values = {
            'c1': request.form.get('col1'),
            'c2': request.form.get('col2'),
            'c3': request.form.get('col3'),
            'c4': request.form.get('col4'),
            'c5': request.form.get('col5')
        }
        for col, value in new_values.items():
            if value is not None:
                setattr(student_data, col, value)
                updated = True
        
        if updated:
            db.session.commit()

    classes = ClassStudent.query.filter_by(code=sector_code,id_admin=user_id).all()
    year_Student = yearStudent.query.filter_by(code=sector_code,id_admin=user_id).all()
    specializations = specializationStudent.query.filter_by(code=sector_code,id_admin=user_id).all()
    grades = GradeStudent.query.filter_by(code=sector_code,id_admin=user_id).all()
    departments = DepartmentStudent.query.filter_by(code=sector_code,id_admin=user_id).all()

    if request.method == 'POST':
        company_name = request.form.get('company_name')
        if company_name:
            if sector_code:
                existing_company = yearStudent.query.filter_by(name=company_name, code=sector_code,id_admin=user_id).first()
                if existing_company:
                    return render_template('student.html',departments=departments,sector_code=sector_code,classes=classes,year_Student=year_Student,grades=grades, specializations=specializations, error="اسم الشركة وكودها موجودان بالفعل")
                else:
                    try:
                        new_company = yearStudent(name=company_name, code=sector_code,id_admin=user_id)
                        db.session.add(new_company)
                        db.session.commit()
                    except exc.IntegrityError as e:
                        db.session.rollback()
                       
                    except Exception as e:
                        db.session.rollback()
                       

       
        specialization_name = request.form.get('specialization_name')
        if specialization_name:
            try:
                existing_job = specializationStudent.query.filter_by(name=specialization_name, code=sector_code,id_admin=user_id).first()
                if not existing_job:
                    new_job = specializationStudent(name=specialization_name, code=sector_code,id_admin=user_id)
                    db.session.add(new_job)
                    db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")
            return redirect(url_for('student'))

        grade_name = request.form.get('grade_name')
        if grade_name:
            if sector_code:
                existing_specialization = GradeStudent.query.filter_by(name=grade_name, code=sector_code,id_admin=user_id).first()

                if existing_specialization:
                    return render_template('student.html',departments=departments,sector_code=sector_code,classes=classes,year_Student=year_Student,grades=grades, specializations=specializations, error="التخصص موجود بالفعل")
                else:
                    try:
                        new_specialization = GradeStudent(name=grade_name, code=sector_code,id_admin=user_id)
                        db.session.add(new_specialization)
                        db.session.commit()
                        return redirect(url_for('student'))
                    except exc.IntegrityError as e:
                        db.session.rollback()
                        return render_template('student.html',departments=departments,sector_code=sector_code,classes=classes, year_Student=year_Student, grades=grades,specializations=specializations, error=f"خطأ في قاعدة البيانات: {e}")
                    except Exception as e:
                        db.session.rollback()
                        return render_template('student.html',departments=departments,sector_code=sector_code,classes=classes, year_Student=year_Student, grades=grades,specializations=specializations, error=f"حدث خطأ: {e}")
        
        class_name = request.form.get('class_name')
        if class_name:
            if sector_code:
                existing_specialization = ClassStudent.query.filter_by(name=class_name, code=sector_code,id_admin=user_id).first()

                if existing_specialization:
                    return render_template('student.html', departments=departments,sector_code=sector_code,year_Student=year_Student,grades=grades, specializations=specializations,classes=classes,error="التخصص موجود بالفعل")
                else:
                    try:
                        new_specialization = ClassStudent(name=class_name, code=sector_code,id_admin=user_id)
                        db.session.add(new_specialization)
                        db.session.commit()
                        return redirect(url_for('student'))
                    except exc.IntegrityError as e:
                        db.session.rollback()
                        return render_template('student.html',departments=departments,sector_code=sector_code,classes=classes, year_Student=year_Student, grades=grades,specializations=specializations, error=f"خطأ في قاعدة البيانات: {e}")
                    except Exception as e:
                        db.session.rollback()
                        return render_template('student.html',departments=departments,sector_code=sector_code,year_Student=year_Student, grades=grades,classes=classes,specializations=specializations, error=f"حدث خطأ: {e}")

        department_name = request.form.get('department_name')
        if department_name:
            if sector_code:
                existing_specialization = DepartmentStudent.query.filter_by(name=department_name, code=sector_code,id_admin=user_id).first()

                if existing_specialization:
                    return render_template('student.html', departments=departments,sector_code=sector_code,year_Student=year_Student,grades=grades, specializations=specializations,classes=classes,error="التخصص موجود بالفعل")
                else:
                    try:
                        new_specialization =DepartmentStudent(name=department_name, code=sector_code,id_admin=user_id)
                        db.session.add(new_specialization)
                        db.session.commit()
                        return redirect(url_for('student'))
                    except exc.IntegrityError as e:
                        db.session.rollback()
                        return render_template('student.html',departments=departments,sector_code=sector_code,classes=classes, year_Student=year_Student, grades=grades,specializations=specializations, error=f"خطأ في قاعدة البيانات: {e}")
                    except Exception as e:
                        db.session.rollback()
                        return render_template('student.html',sector_code=sector_code,departments=departments,year_Student=year_Student, grades=grades,classes=classes,specializations=specializations, error=f"حدث خطأ: {e}")
        
        if 'excel_file' in request.files:
            excel_file = request.files['excel_file']
            if excel_file.filename != '':
                filename = secure_filename(excel_file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                excel_file.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active

                    for row in sheet.iter_rows(min_row=2):
                        name = row[0].value
                        existing_company = yearStudent.query.filter_by(name=name, code=sector_code,id_admin=user_id).first()

                        if not existing_company:
                            new_company = yearStudent(name=name, code=sector_code,id_admin=user_id)
                            db.session.add(new_company)

                    db.session.commit()

                except IntegrityError as e:
                    db.session.rollback()
                    print(f"Database integrity error: {e}")
                except Exception as e:
                    db.session.rollback()
                    print(f"An error occurred: {e}")
                finally:
                    os.remove(filepath)
                return redirect(url_for('student'))

        if 'excel_file_specialization' in request.files:
            specialization_excel_file = request.files['excel_file_specialization']
            if specialization_excel_file.filename != '':
                filename = secure_filename(specialization_excel_file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                specialization_excel_file.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=2):
                        name = row[0].value
                        if not specializationStudent.query.filter_by(name=name, code=sector_code,id_admin=user_id).first():
                            new_specialization = specializationStudent(name=name, code=sector_code,id_admin=user_id)
                            db.session.add(new_specialization)
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    print(f"Error: {e}")
                finally:
                    os.remove(filepath)
                return redirect(url_for('student'))


        if 'excel_file_grade' in request.files:
            excel_file_grade = request.files['excel_file_grade']
            if excel_file_grade.filename != '':
                filename = secure_filename(excel_file_grade.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                excel_file_grade.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active

                    for row in sheet.iter_rows(min_row=2):
                        grade_name = row[0].value  

                        existing_grade = GradeStudent.query.filter_by(name=grade_name, code=sector_code,id_admin=user_id).first()

                        if not existing_grade:
                            new_grade = GradeStudent(name=grade_name, code=sector_code,id_admin=user_id)
                            db.session.add(new_grade)

                    db.session.commit()

                except IntegrityError as e:
                    db.session.rollback()
                    print(f"Database integrity error: {e}")
                except Exception as e:
                    db.session.rollback()
                    print(f"An error occurred: {e}")
                finally:
                    os.remove(filepath)
                return redirect(url_for('student'))

        if 'excel_file_class' in request.files:
            excel_file_class = request.files['excel_file_class']
            if excel_file_class.filename != '':
                filename = secure_filename(excel_file_class.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                excel_file_class.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active

                    for row in sheet.iter_rows(min_row=2):
                        class_name = row[0].value  

                        existing_class = ClassStudent.query.filter_by(name=class_name, code=sector_code,id_admin=user_id).first()

                        if not existing_class:
                            new_class = ClassStudent(name=class_name, code=sector_code,id_admin=user_id)
                            db.session.add(new_class)

                    db.session.commit()

                except IntegrityError as e:
                    db.session.rollback()
                    print(f"Database integrity error: {e}")
                except Exception as e:
                    db.session.rollback()
                    print(f"An error occurred: {e}")
                finally:
                    os.remove(filepath)
                return redirect(url_for('student'))
                
        if 'excel_file_department' in request.files:
            excel_file_department = request.files['excel_file_department']
            if excel_file_department.filename != '':
                filename = secure_filename(excel_file_department.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                excel_file_department.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active

                    for row in sheet.iter_rows(min_row=2):
                        department_name = row[0].value  

                        existing_department = DepartmentStudent.query.filter_by(name=department_name, code=sector_code,id_admin=user_id).first()

                        if not existing_department:
                            new_department = DepartmentStudent(name=department_name, code=sector_code,id_admin=user_id)
                            db.session.add(new_department)

                    db.session.commit()

                except InterruptedError as e:
                    db.session.rollback()
                    print(f"Database integrity error: {e}")
                except Exception as e:
                    db.session.rollback()
                    print(f"An error occurred: {e}")
                finally:
                    os.remove(filepath)
                return redirect(url_for('student'))

    return render_template('student.html',sector_code_display=sector_code_display,student_data=student_data,sector_code=sector_code,departments=departments,year_Student=year_Student, specializations=specializations,grades=grades,classes=classes)
@app.route('/logout')
def logout():
    
    session.clear()
    return redirect(url_for('index'))
@app.route('/delete_department/<int:department_id>', methods=['GET'])
def delete_department(department_id):
    department = DepartmentStudent.query.get(department_id)
    if department:
        db.session.delete(department)
        db.session.commit()
    return redirect(url_for('student'))

@app.route('/delete_department1/<int:department_id>', methods=['GET'])
def delete_department1(department_id):
    department = DepartmentEmployee.query.get(department_id)
    if department:
        db.session.delete(department)
        db.session.commit()
    return redirect(url_for('employee'))

@app.route('/delete_department2/<int:department_id>', methods=['GET'])
def delete_department2(department_id):
    department = DepartmentEmployee1.query.get(department_id)
    if department:
        db.session.delete(department)
        db.session.commit()
    return redirect(url_for('employee'))

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/owner', methods=['GET', 'POST'])
def owner():
    
    sector_id = session.get('sector_id')
    sector_code = session.get('sector_code')
    if not sector_id or not sector_code:
        return redirect(url_for('access')) 
    sector = Sector.query.filter_by(id=sector_id).first()
    unread_notifications = Notification_owner.query.filter_by(company_code=sector_code, viewed=False).all()
    if unread_notifications:
        notification_dot_display = True
    else:
        notification_dot_display = False
    company_names = CompanyName.query.filter_by(code=sector_code).all()
    col_data = ColOrganizer.query.filter_by(code=sector_code).first()
    if request.method == 'POST' and request.form.get('foundation_name') is not None:
        foundation_name = request.form.get('foundation_name')
        address = request.form.get('address')
        phone_number = request.form.get('phone_number')
        state = request.form.get('state')
        country = request.form.get('country')
        description = request.form.get('description')

       
        if len(foundation_name) < 3:
            return "Error: Company name must be at least 3 characters.", 400
        if len(address) < 3:
            return "Error: address must be at least 3 characters.", 400
        if not phone_number.isdigit():
            return "Error: Phone number must contain only digits.", 400
       

        # تحديث البيانات في قاعدة البيانات
        sector.foundation_name = foundation_name
        sector.address = address
        sector.phone_number = phone_number
        sector.state = state
        sector.country = country
        sector.description = description or None

        db.session.commit()
        return redirect(url_for('owner'))

    if request.method == 'POST' and 'profile_image' in request.files:
        profile_image = request.files['profile_image']
        if profile_image and allowed_file(profile_image.filename):
            filename = f"{sector.code}.png" 
            
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            if os.path.exists(file_path):
                os.remove(file_path)
           
            profile_image.save(file_path)
            
          
            sector.image_path = f'{filename}'
            db.session.commit()
            
            return redirect(url_for('owner'))

    if not col_data:
        col_data = ColOrganizer(code=sector_code, c1='', c2='', c3='', c4='', c5='')
        db.session.add(col_data)
        db.session.commit()
    if request.method == 'POST':
        col_data.c1 = request.form.get('column1', col_data.c1)
        col_data.c2 = request.form.get('column2', col_data.c2)
        col_data.c3 = request.form.get('column3', col_data.c3)
        col_data.c4 = request.form.get('column4', col_data.c4)
        col_data.c5 = request.form.get('column5', col_data.c5)
        db.session.commit()
        return redirect(url_for('owner'))
    if not sector_id:
        return redirect(url_for('signup_sector')) 
    
    
    all_sectors = organizer2.query.filter_by(code=sector_code).all()
    visible_columns = ColumnPreference.query.filter_by(code=sector_code, visible=True).all()
    visible_column_names = [col.column_name for col in visible_columns]
    if not sector:
        return redirect(url_for('signup_sector'))  
    sector_id = session.get('sector_id')
    sector_code = session.get('sector_code')
    if request.method == 'POST' and request.form.get('company_name')!=None:
        company_name = request.form.get('company_name')

        if company_name and sector_code:
            existing_company = CompanyName.query.filter_by(name=company_name, code=sector_code).first()
            if existing_company:
                return render_template('owner.html',sector_code=sector_code, sector=sector, all_sectors=all_sectors, company_names=company_names, error="اسم الشركة وكودها موجودان بالفعل")
            else:
                try:
                    new_company = CompanyName(name=company_name, code=sector_code)
                    db.session.add(new_company)
                    db.session.commit()
                    return redirect(url_for('owner')) 
                except exc.IntegrityError as e:
                    db.session.rollback()
                    return render_template('owner.html',sector_code=sector_code, sector=sector, all_sectors=all_sectors, company_names=company_names, error=f"خطأ في قاعدة البيانات: {e}")
                except Exception as e:
                    db.session.rollback()
                    return render_template('owner.html',sector_code=sector_code, sector=sector, all_sectors=all_sectors, company_names=company_names, error=f"حدث خطأ: {e}")

    

    if 'excel_file' in request.files:
        excel_file = request.files['excel_file']
        if excel_file.filename != '':
            filename = secure_filename(excel_file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            excel_file.save(filepath)

            try:
                workbook = load_workbook(filepath)
                sheet = workbook.active  

                for row in sheet.iter_rows(min_row=2):
                    name = row[0].value
                    print(name) 
                    print(sector_code)

                    existing_company = CompanyName.query.filter_by(name=name, code=sector_code).first()

                    if existing_company:
                        print(f"Duplicate entry detected: {name} with code {sector_code}")
                        continue 

                    new_company = CompanyName(name=name, code=sector_code)
                    db.session.add(new_company)

                db.session.commit()
            except exc.IntegrityError as e:
                db.session.rollback()
                print(f"Database integrity error: {e}")
            except Exception as e:
                db.session.rollback()
                print(f"An error occurred: {e}")
                print(f"Error processing Excel file: {e}")
                return render_template('owner.html',col_data=col_data,sector_code=sector_code, sector=sector, all_sectors=all_sectors, error=f"Error processing Excel file: {e}")
            finally:
                os.remove(filepath) 

    return render_template('owner.html',notification_dot_display=notification_dot_display,col_data=col_data,sector_code=sector_code, sector=sector, all_sectors=all_sectors, company_names=company_names)



@app.route('/show_org', methods=['GET', 'POST'])
def show_org():
    organizer_id = session.get('organizer_id')
    sector_id = session.get('sector_id')
    sector_code = session.get('sector_code')
    if not sector_id or not sector_code:
        return redirect(url_for('access')) 
    sector = Sector.query.filter_by(id=sector_id).first()

    company_names = CompanyName.query.filter_by(code=sector_code).all()
    col_data = ColOrganizer.query.filter_by(code=sector_code).first()
    all_sectors = organizer2.query.filter_by(code=sector_code).all()
    visible_columns = ColumnPreference.query.filter_by(code=sector_code, visible=True).all()
    visible_column_names = [col.column_name for col in visible_columns]
 
    return render_template('show_org.html',col_data=col_data,sector_code=sector_code, sector=sector, all_sectors=all_sectors, company_names=company_names,organizer_id=organizer_id,sector_id=sector_id)

@app.route('/success')
def success():
    username = request.args.get('username')
    email = "user@example.com"  

    preferences = ColumnPreference.query.filter_by(email=email).all()
    visible_columns = {pref.column_name: pref.visible for pref in preferences}

    personal_data = Personal.query.all()
    return render_template('success.html', 
                           username=username, 
                           personal_data=personal_data, 
                           visible_columns=visible_columns)


@app.route('/manager', methods=['GET', 'POST'])
def manager():
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    if not sector_code or not (organizer_id or sector_id ):
        return redirect(url_for('access'))
    employee_name = ColEmployee.query.filter_by(code=sector_code).first()
    employees = Employee6.query.filter_by(code=sector_code).all()
    return render_template('manager.html', employees=employees,employee_name=employee_name,sector_code=sector_code,organizer_id=organizer_id,sector_id=sector_id)


@app.route('/manager_adm', methods=['GET', 'POST'])
def manager_adm():
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    if not sector_code or not ( user_id):
        return redirect(url_for('access'))
    employee_name = ColEmployee.query.filter_by(code=sector_code,id_admin=user_id,).first()
    employees = Employee6.query.filter_by(code=sector_code,id_admin=user_id).all()
    return render_template('manager_adm.html', employees=employees,employee_name=employee_name,sector_code=sector_code)

@app.route('/managertomanger_adm', methods=['GET', 'POST'])
def managertomanger_adm():
    all_data = head.query.all()  
    for data in all_data:
        print(f"ID: {data.id}, Username: {data.username}, Email: {data.email}, Phone: {data.phone}")
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    if not sector_code or not (user_id):
        return redirect(url_for('access'))
    employee_name = ColEmployee.query.filter_by(code=sector_code,id_admin=user_id).first()
    managers = ManagerEmployee.query.filter_by(code=sector_code,id_admin=user_id).all() 
    return render_template('managertomanger_adm.html', managers=managers,sector_code=sector_code,employee_name=employee_name)


@app.route('/assign_manager1_adm', methods=['POST'])
def assign_manager1_adm():
    if request.method == 'POST':

        manager_id = request.form.get('manager')  
    
        employee_ids = request.form.getlist('employees')  

        if manager_id and employee_ids:
            manager = ManagerEmployee.query.get(manager_id)  
            if manager:
                manager_to_update = Employee6.query.filter_by(email=manager.email).first()
                if manager_to_update:
                    manager_to_update.is_head = "ttt"
                    db.session.commit()
                
                for emp_id in employee_ids:
                    employee = ManagerEmployee.query.get(emp_id)
                    if employee:

    
                        employee.manager_id = manager.id
                        employee.manager_name = manager.username
                        db.session.commit()

                existing_manager = head.query.filter_by(email=manager.email).first()
                if not existing_manager:
                    manager_employee = head(
                        id_admin=manager.id_admin,
                        code=manager.code,
                        username=manager.username,
                        email=manager.email,
                        password=manager.password,
                        birthdate=manager.birthdate,
                        stage=manager.stage,
                        university=manager.university,
                        major=manager.major,
                        phone=manager.phone,
                        country=manager.country,
                        province=manager.province,
                        col1=manager.col1,
                        col2=manager.col2,
                        col3=manager.col3,
                        col4=manager.col4,
                        col5=manager.col5,
                        branch=manager.branch
                    )
                    db.session.add(manager_employee)  
                    db.session.commit() 

                return redirect('/managertomanger_adm') 
    return redirect(url_for('managertomanger_adm'))

@app.route('/assign_manager_adm', methods=['GET', 'POST'])
def assign_manager_adm():
    if request.method == 'POST':
        manager_id = request.form.get('manager')  
        employee_ids = request.form.getlist('employees')  
        
        if manager_id and employee_ids:
            manager = Employee6.query.get(manager_id) 
            if manager:
                if not manager.is_manager:  
                    manager.is_manager = True
                    db.session.commit()
                for emp_id in employee_ids:
                    employee = Employee6.query.get(emp_id)
                    if employee:
                  
                        employee.manager_id = manager.id
                        employee.manager_name = manager.username
                    
                        db.session.commit()

                existing_manager = ManagerEmployee.query.filter_by(email=manager.email).first()
                print(manager.id_admin)
                print(manager.id_admin)
                print(manager.id_admin)
                if not existing_manager:
                    manager_employee = ManagerEmployee(
                        id_admin=manager.id_admin,
                        code=manager.code,
                        username=manager.username,
                        email=manager.email,
                        password=manager.password,
                        birthdate=manager.birthdate,
                        stage=manager.stage,
                        university=manager.university,
                        major=manager.major,
                        phone=manager.phone,
                        country=manager.country,
                        province=manager.province,
                        col1=manager.col1,
                        col2=manager.col2,
                        col3=manager.col3,
                        col4=manager.col4,
                        col5=manager.col5,
                        branch=manager.branch
                    )
                    db.session.add(manager_employee)  
                    db.session.commit() 

                return redirect('/manager_adm') 
            else:
                return "Manager not found", 404 
        else:
            return "Error: Manager or Employees not selected", 400  
    
    return render_template('manager_adm.html')

@app.route('/show_stu', methods=['GET', 'POST'])
def show_stu():
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    if not sector_code or not (organizer_id or sector_id):
        return redirect(url_for('access'))
    employees = SectorStudent8.query.filter_by(code=sector_code).all()
    return render_template('show_stu.html', employees=employees,sector_code=sector_code,organizer_id=organizer_id,sector_id=sector_id)

@app.route('/delete_stu/<int:sector_id>', methods=['GET'])
def delete_stu(sector_id):
    sector = SectorStudent8.query.get_or_404(sector_id)
    db.session.delete(sector)
    db.session.commit()
    return redirect(url_for('show_stu'))

@app.route('/show_adm', methods=['GET', 'POST'])
def show_adm():
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    if not sector_code or not (organizer_id or sector_id):
        return redirect(url_for('access'))
    employees = administrator2.query.filter_by(code=sector_code).all()
    admin_data = ColOrganizer.query.filter_by(code=sector_code).first()
    column_names = ColAdministrator.query.filter_by(code=sector_code).first()
    return render_template('show_adm.html', employees=employees,sector_code=sector_code,columns=column_names,sector_id=sector_id,organizer_id=organizer_id,
    c1=admin_data.c1,
    c2=admin_data.c2,
    c3=admin_data.c3,
    c4=admin_data.c4,
    c5=admin_data.c5)

@app.route('/delete_adm/<int:sector_id>', methods=['GET'])
def delete_adm(sector_id):
    sector = administrator2.query.get_or_404(sector_id)
    db.session.delete(sector)
    db.session.commit()
    return redirect(url_for('show_adm'))

@app.route('/show_emp', methods=['GET', 'POST'])
def show_emp():
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    if not sector_code or not (organizer_id or sector_id):
        return redirect(url_for('access'))
    employee_name = ColEmployee.query.filter_by(code=sector_code).first()
    employees = Employee6.query.filter_by(code=sector_code).all()
    return render_template('show_emp.html', employees=employees,employee_name=employee_name,sector_code=sector_code,organizer_id=organizer_id,sector_id=sector_id)

@app.route('/delete_emp/<int:sector_id>', methods=['GET'])
def delete_emp(sector_id):
    sector = Employee6.query.get_or_404(sector_id)
    db.session.delete(sector)
    db.session.commit()
    return redirect(url_for('show_emp'))

@app.route('/managertomanger', methods=['GET', 'POST'])
def managertomanger():
    all_data = head.query.all()  
    for data in all_data:
        print(f"ID: {data.id}, Username: {data.username}, Email: {data.email}, Phone: {data.phone}")
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    if not sector_code or not (organizer_id or sector_id or user_id):
        return redirect(url_for('access'))
    employee_name = ColEmployee.query.filter_by(code=sector_code).first()
    managers = ManagerEmployee.query.filter_by(code=sector_code).all() 
    return render_template('managertomanger.html', managers=managers,sector_code=sector_code,employee_name=employee_name,organizer_id=organizer_id,sector_id=sector_id)

@app.route('/assign_manager1', methods=['POST'])
def assign_manager1():
    if request.method == 'POST':
        manager_id = request.form.get('manager')  
        employee_ids = request.form.getlist('employees')  

        if manager_id and employee_ids:
            manager = ManagerEmployee.query.get(manager_id) 
            if manager:
                manager_to_update = Employee6.query.filter_by(email=manager.email).first()
                if manager_to_update:
                    manager_to_update.is_head = "ttt"
                    db.session.commit()

                for emp_id in employee_ids:
                    employee = ManagerEmployee.query.get(emp_id)
                    if employee:
                        employee.manager_id = manager.id
                        employee.manager_name = manager.username
                        db.session.commit()

                # التحقق من وجود المدير في جدول head
                existing_manager = head.query.filter_by(email=manager.email).first()
                if not existing_manager:
                    manager_employee = head(
                        id_admin=manager.id_admin,
                        code=manager.code,
                        username=manager.username,
                        email=manager.email,
                        password=manager.password,
                        birthdate=manager.birthdate,
                        stage=manager.stage,
                        university=manager.university,
                        major=manager.major,
                        phone=manager.phone,
                        country=manager.country,
                        province=manager.province,
                        col1=manager.col1,
                        col2=manager.col2,
                        col3=manager.col3,
                        col4=manager.col4,
                        col5=manager.col5,
                        branch=manager.branch
                    )
                    db.session.add(manager_employee)  
                    db.session.commit() 

                return redirect('/managertomanger') 
    
    return redirect(url_for('managertomanger'))

@app.route('/assign_manager', methods=['GET', 'POST'])
def assign_manager():
    if request.method == 'POST':
        manager_id = request.form.get('manager')  
        employee_ids = request.form.getlist('employees')  
        
        if manager_id and employee_ids:
            manager = Employee6.query.get(manager_id) 
            if manager:
                # تعيين الموظف كمدير (is_manager = True)
                if not manager.is_manager:  # إذا لم يكن المدير الحالي بالفعل مديرًا
                    manager.is_manager = True
                    db.session.commit()

                for emp_id in employee_ids:
                    employee = Employee6.query.get(emp_id)
                    if employee:
                        # تعيين المدير للموظف
                        employee.manager_id = manager.id
                        employee.manager_name = manager.username
                        db.session.commit()

                # نقل المدير إلى جدول ManagerEmployee إذا لم يكن موجودًا بالفعل
                existing_manager = ManagerEmployee.query.filter_by(email=manager.email).first()
                if not existing_manager:
                    manager_employee = ManagerEmployee(
                        id_admin=manager.id_admin,
                        code=manager.code,
                        username=manager.username,
                        email=manager.email,
                        password=manager.password,
                        birthdate=manager.birthdate,
                        stage=manager.stage,
                        university=manager.university,
                        major=manager.major,
                        phone=manager.phone,
                        country=manager.country,
                        province=manager.province,
                        col1=manager.col1,
                        col2=manager.col2,
                        col3=manager.col3,
                        col4=manager.col4,
                        col5=manager.col5,
                        branch=manager.branch
                    )
                    db.session.add(manager_employee)  
                    db.session.commit() 

                return redirect('/manager') 
            else:
                return "Manager not found", 404 
        else:
            return "Error: Manager or Employees not selected", 400  
    
    return render_template('manager.html')




if __name__ == '__main__':
    port = int(os.environ.get("PORT", 4000))
    app.run(host="0.0.0.0", port=port)